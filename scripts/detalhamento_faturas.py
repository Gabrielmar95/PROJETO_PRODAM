"""
detalhamento_faturas.py — relatorio granular de faturas por devedor.

Gera breakdown completo por categoria para cada devedor do portfolio:
  - Por situacao (Emitida / Parc. Paga / EmPerdas / Totalmente Paga / Cancelada)
  - Por prescricao (Prescrita / Nao-Prescrita — regra 5 anos do Art. 206 Sec.5 I CC)
  - Por exigibilidade (cruzamento situacao x prescricao x cadeia_completude)

Outputs em C:\\Users\\gabri\\Desktop\\PROJETO_PRODAM\\DETALHAMENTO_FATURAS\\:
  - DETALHAMENTO_FATURAS.xlsx         (4 abas: CONSOLIDADO, POR_FATURA, RESUMO_GLOBAL, SEM_MATCH_SSOT)
  - DETALHAMENTO_FATURAS.html         (dashboard interativo)
  - RESUMO_DETALHAMENTO.md            (texto)

Matching exato de cliente via dicionario ALIAS (evita contaminacao SEMA x SEMAD-MANAUS).
"""

from __future__ import annotations
import sys, json, sqlite3
from pathlib import Path
from datetime import date
from decimal import Decimal
from collections import defaultdict

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent
DB = ROOT / "prodam.db"
PROFILES = ROOT / "PRODAM_DOCS" / "profiles.json"
OUT = ROOT / "DETALHAMENTO_FATURAS"
OUT.mkdir(exist_ok=True)

# Cutoff prescricao: hoje - 5 anos (Art. 206 Sec.5 I CC - divida liquida)
HOJE = date.today()
CUTOFF_ANO = HOJE.year - 5
CUTOFF_MES = HOJE.month

# Mapa EXATO sigla SSOT -> clientes no DB. Resolve ambiguidades manualmente.
# (Se nao listado aqui, usa match exato por nome entre sigla e cliente.)
ALIAS = {
    "SES/SUSAM": ["SES"],
    "FAAR/SEDEL": ["SEDEL", "FAAR - EXTINTA LEI 6.225 INCP PELA SEDEL"],
    "FMT": ["FMT-HVD"],
    "POLICIA CIVIL": ["POLICIA CIVIL"],
    "POLÍCIA CIVIL": ["POLICIA CIVIL"],
    "SALUX": ["SALUX INFORMATIZAÇÃO EM SAUDE S/A"],
    "BRADESCO": ["BANCO BRADESCO S. A.", "BRADESCO FINANCIAMENTO"],
    "BANCO_BMG": ["BANCO BMG"],
    "BANCO_SAFRA": ["BANCO SAFRA"],
    "BANCO_MASTER": ["BANCO MASTER S/A"],
    "BANCO_DAYCOVAL": ["BANCO DAYCOVAL"],
    "BANCO_DO_BRASIL": ["BANCO DO BRASIL - SETOR PUBL."],
    "BANCO_SICOOB": ["BANCO SICOOB"],
    "B23": ["B23 TECNOLOGIA E PAGAMENTOS LTDA"],
    "FENIXSOFT": ["FENIXSOFT", "FENIXSOFT2"],
    "EYES": ["EYES NWHERE SISTEMAS INTELIGENTES DE IMAGEM LTDA"],
    "THOMAS_GREG": ["THOMAS GREG E SONS"],
    "SULAMERICA": ["SUL AMERICA"],
    "PROVER": ["PROVER PROMOÇÃO DE VENDAS LTDA."],
    "PRONTO_PAGUEI": ["PRONTO PAGUEI GESTÃO FINANCEIRA LTDA"],
    "PARCELAMOS_TUDO": ["PARCELAMOS TUDO PONTOCOM SOLUÇÕES EM PAGAMENTO LTDA"],
    "ROCHA_E_ROCHA": ["ROCHA E ROCHA SERVICOS DE CORRETAGEM DE IMOVEIS LTDA"],
    "AM_CARD": ["AM CARD - ADMINISTRADORA DE CARTÕES S/S LTDA"],
    "AMETRAN": ["AMETRAN"],
    "EASYTECH": ["EASYTECH"],
    "PADRAO_SYSTEM": ["PADRAO SYSTEM"],
    "CAPEMISA": ["CAPEMISA"],
    "CASPEB": ["CASPEB"],
    "ASSALE": ["ASSALE"],
    "UGPADEAM": ["UGPADEAM"],
    "UGPE": ["UGPE - UNIDADE GESTORA DE PROJETOS ESPECIAIS"],
    "FUNJEAM": ["FUNJEAM-TJ"],
    "FUNJEAM-TJ": ["FUNJEAM-TJ"],
    "ANOREG": ["ANOREG/AM"],
    "CNB": ["CNB - SEÇÃO AMAZONAS"],
    "AGIR": ["ASSOCIAÇÃO DE GESTÃO INOVAÇÃO E RESULTADOS EM SAUDE"],
    "Moraes_e_Almeida": ["Moraes e Almeida Ltda"],
}


def norm(s: str) -> str:
    return (
        (s or "").upper().strip().replace("/", "_").replace("-", "_").replace(" ", "_")
    )


def resolver_clientes_db(sigla: str, clientes_db: set) -> list:
    """Dada uma sigla do SSOT, retorna lista de strings exatas no DB."""
    if sigla in ALIAS:
        return [c for c in ALIAS[sigla] if c in clientes_db]
    # Tentativa 1: match exato
    if sigla in clientes_db:
        return [sigla]
    # Tentativa 2: match exato ignorando case/acento
    sigla_n = norm(sigla)
    candidatos = [c for c in clientes_db if norm(c) == sigla_n]
    return candidatos


def is_prescrita(competencia: str) -> bool | None:
    """Parseia 'MM/AAAA' e retorna True se prescrita (5 anos regra)."""
    if not competencia or "/" not in competencia:
        return None
    try:
        mes, ano = competencia.split("/")
        mes, ano = int(mes), int(ano)
    except (ValueError, IndexError):
        return None
    if ano < CUTOFF_ANO:
        return True
    if ano == CUTOFF_ANO and mes < CUTOFF_MES:
        return True
    return False


def classificar_exigibilidade(
    situacao: str, prescrita: bool | None, cadeia: str | None
) -> str:
    """Classifica exigibilidade processual combinando 3 dimensoes."""
    if situacao == "Cancelada":
        return "CANCELADA"
    if situacao == "Totalmente Paga":
        return "QUITADA"
    if prescrita is True:
        return "PRESCRITA_NAO_EXIGIVEL"
    if prescrita is None:
        return "INDETERMINADA_SEM_DATA"
    # Nao prescrita + em aberto/parcial/perdas
    if cadeia in ("COMPLETA", "FORTE"):
        return "EXIGIVEL_FORTE"
    if cadeia == "MEDIA":
        return "EXIGIVEL_MEDIA"
    if cadeia == "FRACA":
        return "EXIGIVEL_FRACA"
    return "EXIGIVEL_SEM_CADEIA"


def main():
    print(f"[1/5] Carregando DB e SSOT...")
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    clientes_db = {
        r[0] for r in conn.execute("SELECT DISTINCT cliente FROM spcf_faturas") if r[0]
    }
    print(f"      DB: {len(clientes_db)} clientes distintos")

    profiles = json.loads(PROFILES.read_text(encoding="utf-8"))
    siglas = [k for k in profiles.keys() if not k.startswith("_")]
    print(f"      SSOT: {len(siglas)} siglas")

    print(f"[2/5] Classificando 1927 faturas...")
    faturas = []
    for r in conn.execute("""
        SELECT id, nf, cliente, contrato_num, valor_bruto, competencia,
               situacao, cadeia_completude, tem_pdf, total_empenhos_vinculados
        FROM spcf_faturas
    """):
        d = dict(r)
        d["prescrita"] = is_prescrita(d["competencia"])
        d["exigibilidade"] = classificar_exigibilidade(
            d["situacao"], d["prescrita"], d["cadeia_completude"]
        )
        faturas.append(d)

    print(f"[3/5] Consolidando por devedor ({len(siglas)} siglas)...")
    consolidado = []
    sem_match = []  # siglas SSOT que nao tem clientes no DB
    clientes_usados = set()

    for sigla in siglas:
        clientes = resolver_clientes_db(sigla, clientes_db)
        if not clientes:
            sem_match.append(sigla)
            continue
        clientes_usados.update(clientes)
        faturas_dev = [f for f in faturas if f["cliente"] in clientes]

        # Breakdown por situacao
        por_situacao = defaultdict(lambda: {"qtd": 0, "val": Decimal(0)})
        for f in faturas_dev:
            s = f["situacao"] or "SEM_SITUACAO"
            por_situacao[s]["qtd"] += 1
            por_situacao[s]["val"] += Decimal(str(f["valor_bruto"] or 0))

        # Breakdown por prescricao
        prescritas = [f for f in faturas_dev if f["prescrita"] is True]
        nao_prescritas = [f for f in faturas_dev if f["prescrita"] is False]
        sem_data = [f for f in faturas_dev if f["prescrita"] is None]

        # Breakdown por exigibilidade
        por_exig = defaultdict(lambda: {"qtd": 0, "val": Decimal(0)})
        for f in faturas_dev:
            por_exig[f["exigibilidade"]]["qtd"] += 1
            por_exig[f["exigibilidade"]]["val"] += Decimal(str(f["valor_bruto"] or 0))

        p = profiles[sigla]
        consolidado.append(
            {
                "sigla": sigla,
                "categoria": p.get("categoria", "?"),
                "clientes_db": ", ".join(clientes),
                "total_qtd": len(faturas_dev),
                "total_val": float(
                    sum(Decimal(str(f["valor_bruto"] or 0)) for f in faturas_dev)
                ),
                # Por situacao
                "emitida_qtd": por_situacao["Emitida"]["qtd"],
                "emitida_val": float(por_situacao["Emitida"]["val"]),
                "parcial_qtd": por_situacao["Parcialmente Paga"]["qtd"],
                "parcial_val": float(por_situacao["Parcialmente Paga"]["val"]),
                "perdas_qtd": por_situacao["EmPerdas"]["qtd"],
                "perdas_val": float(por_situacao["EmPerdas"]["val"]),
                "quitada_qtd": por_situacao["Totalmente Paga"]["qtd"],
                "quitada_val": float(por_situacao["Totalmente Paga"]["val"]),
                "cancelada_qtd": por_situacao["Cancelada"]["qtd"],
                "cancelada_val": float(por_situacao["Cancelada"]["val"]),
                # Por prescricao
                "prescritas_qtd": len(prescritas),
                "prescritas_val": float(
                    sum(Decimal(str(f["valor_bruto"] or 0)) for f in prescritas)
                ),
                "nao_prescritas_qtd": len(nao_prescritas),
                "nao_prescritas_val": float(
                    sum(Decimal(str(f["valor_bruto"] or 0)) for f in nao_prescritas)
                ),
                "sem_data_qtd": len(sem_data),
                # Por exigibilidade
                "exig_forte_qtd": por_exig["EXIGIVEL_FORTE"]["qtd"],
                "exig_forte_val": float(por_exig["EXIGIVEL_FORTE"]["val"]),
                "exig_media_qtd": por_exig["EXIGIVEL_MEDIA"]["qtd"],
                "exig_media_val": float(por_exig["EXIGIVEL_MEDIA"]["val"]),
                "exig_fraca_qtd": por_exig["EXIGIVEL_FRACA"]["qtd"],
                "exig_fraca_val": float(por_exig["EXIGIVEL_FRACA"]["val"]),
                "prescrita_qtd": por_exig["PRESCRITA_NAO_EXIGIVEL"]["qtd"],
                "prescrita_val": float(por_exig["PRESCRITA_NAO_EXIGIVEL"]["val"]),
                # SSOT comparativo
                "ssot_val_exig": float(p.get("val_exig") or 0),
                "ssot_faturas_exigiveis": p.get("faturas_exigiveis") or 0,
                "ssot_faturas_total": p.get("faturas_total") or 0,
            }
        )

    clientes_db_sem_sigla = sorted(clientes_db - clientes_usados)

    print(f"      {len(consolidado)} devedores consolidados")
    print(f"      {len(sem_match)} siglas SSOT sem match no DB")
    print(f"      {len(clientes_db_sem_sigla)} clientes DB fora do portfolio")

    print(f"[4/5] Gerando Excel...")
    gerar_xlsx(consolidado, faturas, sem_match, clientes_db_sem_sigla)

    print(f"[5/5] Gerando HTML + MD...")
    gerar_html(consolidado)
    gerar_md(consolidado, sem_match, clientes_db_sem_sigla)

    print(f"\nSaida: {OUT}")
    print(f"  - DETALHAMENTO_FATURAS.xlsx")
    print(f"  - DETALHAMENTO_FATURAS.html")
    print(f"  - RESUMO_DETALHAMENTO.md")
    print(f"\nTOTAIS GLOBAIS:")
    print(f"  Faturas total DB:       {len(faturas):>5}")
    print(f"  Em devedores portfolio: {sum(c['total_qtd'] for c in consolidado):>5}")
    print(
        f"  Prescritas:             {sum(c['prescritas_qtd'] for c in consolidado):>5}  R$ {sum(c['prescritas_val'] for c in consolidado):>15,.2f}"
    )
    print(
        f"  Nao-prescritas:         {sum(c['nao_prescritas_qtd'] for c in consolidado):>5}  R$ {sum(c['nao_prescritas_val'] for c in consolidado):>15,.2f}"
    )
    print(
        f"  Exig. FORTE:            {sum(c['exig_forte_qtd'] for c in consolidado):>5}  R$ {sum(c['exig_forte_val'] for c in consolidado):>15,.2f}"
    )
    print(
        f"  Parcialmente Pagas:     {sum(c['parcial_qtd'] for c in consolidado):>5}  R$ {sum(c['parcial_val'] for c in consolidado):>15,.2f}"
    )
    print(
        f"  Quitadas:               {sum(c['quitada_qtd'] for c in consolidado):>5}  R$ {sum(c['quitada_val'] for c in consolidado):>15,.2f}"
    )


def gerar_xlsx(consolidado, faturas, sem_match, clientes_orfaos):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("  [WARN] openpyxl nao instalado, pulando Excel")
        return

    wb = Workbook()

    # Aba 1: CONSOLIDADO

    ws.title = "CONSOLIDADO"
    headers = [
        "Sigla",
        "Categoria",
        "Clientes DB",
        "Total Qtd",
        "Total Valor",
        "Emitida Qtd",
        "Emitida Val",
        "Parcial Qtd",
        "Parcial Val",
        "EmPerdas Qtd",
        "EmPerdas Val",
        "Quitada Qtd",
        "Quitada Val",
        "Cancelada Qtd",
        "Cancelada Val",
        "Prescritas Qtd",
        "Prescritas Val",
        "Nao-Presc Qtd",
        "Nao-Presc Val",
        "Exig FORTE Qtd",
        "Exig FORTE Val",
        "Exig MEDIA Qtd",
        "Exig MEDIA Val",
        "Exig FRACA Qtd",
        "Exig FRACA Val",
        "SSOT val_exig",
        "SSOT faturas_exig",
        "SSOT faturas_total",
    ]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=1, column=col).fill = PatternFill("solid", fgColor="1F3864")

    for c in sorted(consolidado, key=lambda x: -x["total_val"]):
        ws.append(
            [
                c["sigla"],
                c["categoria"],
                c["clientes_db"],
                c["total_qtd"],
                c["total_val"],
                c["emitida_qtd"],
                c["emitida_val"],
                c["parcial_qtd"],
                c["parcial_val"],
                c["perdas_qtd"],
                c["perdas_val"],
                c["quitada_qtd"],
                c["quitada_val"],
                c["cancelada_qtd"],
                c["cancelada_val"],
                c["prescritas_qtd"],
                c["prescritas_val"],
                c["nao_prescritas_qtd"],
                c["nao_prescritas_val"],
                c["exig_forte_qtd"],
                c["exig_forte_val"],
                c["exig_media_qtd"],
                c["exig_media_val"],
                c["exig_fraca_qtd"],
                c["exig_fraca_val"],
                c["ssot_val_exig"],
                c["ssot_faturas_exigiveis"],
                c["ssot_faturas_total"],
            ]
        )

    # Aba 2: POR_FATURA
    ws2 = wb.create_sheet("POR_FATURA")
    ws2.append(
        [
            "ID",
            "NF",
            "Cliente",
            "Contrato",
            "Valor",
            "Competencia",
            "Situacao",
            "Prescrita",
            "Cadeia",
            "Exigibilidade",
            "Tem PDF",
        ]
    )
    for col in range(1, 12):
        ws2.cell(row=1, column=col).font = Font(bold=True, color="FFFFFF")
        ws2.cell(row=1, column=col).fill = PatternFill("solid", fgColor="1F3864")
    for f in faturas:
        ws2.append(
            [
                f["id"],
                f["nf"],
                f["cliente"],
                f["contrato_num"],
                f["valor_bruto"],
                f["competencia"],
                f["situacao"],
                (
                    "SIM"
                    if f["prescrita"]
                    else ("NAO" if f["prescrita"] is False else "?")
                ),
                f["cadeia_completude"],
                f["exigibilidade"],
                "SIM" if f["tem_pdf"] else "NAO",
            ]
        )

    # Aba 3: RESUMO_GLOBAL
    ws3 = wb.create_sheet("RESUMO_GLOBAL")
    ws3.append(["Metrica", "Qtd", "Valor"])
    ws3["A1"].font = Font(bold=True)
    ws3.append(
        [
            "Total faturas DB",
            len(faturas),
            sum(float(f["valor_bruto"] or 0) for f in faturas),
        ]
    )
    ws3.append(
        [
            "Em devedores portfolio",
            sum(c["total_qtd"] for c in consolidado),
            sum(c["total_val"] for c in consolidado),
        ]
    )
    ws3.append(["--- POR SITUACAO ---", "", ""])
    for sit_key, sit_label in [
        ("emitida", "Emitida"),
        ("parcial", "Parcialmente Paga"),
        ("perdas", "EmPerdas"),
        ("quitada", "Totalmente Paga"),
        ("cancelada", "Cancelada"),
    ]:
        ws3.append(
            [
                sit_label,
                sum(c[f"{sit_key}_qtd"] for c in consolidado),
                sum(c[f"{sit_key}_val"] for c in consolidado),
            ]
        )
    ws3.append(["--- POR PRESCRICAO ---", "", ""])
    ws3.append(
        [
            "Prescritas (5+ anos)",
            sum(c["prescritas_qtd"] for c in consolidado),
            sum(c["prescritas_val"] for c in consolidado),
        ]
    )
    ws3.append(
        [
            "Nao-prescritas",
            sum(c["nao_prescritas_qtd"] for c in consolidado),
            sum(c["nao_prescritas_val"] for c in consolidado),
        ]
    )
    ws3.append(["--- POR EXIGIBILIDADE ---", "", ""])
    for ex_key, ex_label in [
        ("exig_forte", "Exigivel FORTE"),
        ("exig_media", "Exigivel MEDIA"),
        ("exig_fraca", "Exigivel FRACA"),
    ]:
        ws3.append(
            [
                ex_label,
                sum(c[f"{ex_key}_qtd"] for c in consolidado),
                sum(c[f"{ex_key}_val"] for c in consolidado),
            ]
        )

    # Aba 4: SEM_MATCH
    ws4 = wb.create_sheet("SEM_MATCH_SSOT")
    ws4.append(["Tipo", "Valor"])
    ws4.append(["Siglas SSOT sem cliente no DB:", ""])
    for s in sem_match:
        ws4.append(["", s])
    ws4.append(["", ""])
    ws4.append(
        ["Clientes no DB fora do portfolio (devedores nao listados no SSOT):", ""]
    )
    for c in clientes_orfaos:
        ws4.append(["", c])

    # Ajustar largura
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            sheet.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    wb.save(OUT / "DETALHAMENTO_FATURAS.xlsx")


def gerar_html(consolidado):
    # Dashboard com filtros e tabela
    data_json = json.dumps(
        [
            {
                "sigla": c["sigla"],
                "categoria": c["categoria"],
                "total_qtd": c["total_qtd"],
                "total_val": c["total_val"],
                "emitida_qtd": c["emitida_qtd"],
                "emitida_val": c["emitida_val"],
                "parcial_qtd": c["parcial_qtd"],
                "parcial_val": c["parcial_val"],
                "perdas_qtd": c["perdas_qtd"],
                "perdas_val": c["perdas_val"],
                "quitada_qtd": c["quitada_qtd"],
                "quitada_val": c["quitada_val"],
                "cancelada_qtd": c["cancelada_qtd"],
                "cancelada_val": c["cancelada_val"],
                "prescritas_qtd": c["prescritas_qtd"],
                "prescritas_val": c["prescritas_val"],
                "nao_prescritas_qtd": c["nao_prescritas_qtd"],
                "nao_prescritas_val": c["nao_prescritas_val"],
                "exig_forte_qtd": c["exig_forte_qtd"],
                "exig_forte_val": c["exig_forte_val"],
                "exig_media_qtd": c["exig_media_qtd"],
                "exig_media_val": c["exig_media_val"],
                "exig_fraca_qtd": c["exig_fraca_qtd"],
                "exig_fraca_val": c["exig_fraca_val"],
            }
            for c in sorted(consolidado, key=lambda x: -x["total_val"])
        ]
    )

    html = (
        """<!DOCTYPE html><html lang="pt-BR"><head><meta charset="UTF-8">
<title>Detalhamento de Faturas — PRODAM</title>
<style>
body { font-family: -apple-system, Segoe UI, sans-serif; margin: 0; padding: 20px; background: #F5F5F5; }
h1 { color: #1F3864; border-bottom: 3px solid #B8963E; padding-bottom: 10px; }
.kpis { display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr)); gap: 15px; margin: 20px 0; }
.kpi { background: white; padding: 15px; border-left: 4px solid #1F3864; border-radius: 4px; }
.kpi-label { font-size: 11px; text-transform: uppercase; color: #666; }
.kpi-val { font-size: 24px; font-weight: bold; color: #1F3864; }
.kpi-sub { font-size: 11px; color: #888; margin-top: 4px; }
.filters { background: white; padding: 15px; margin: 15px 0; border-radius: 4px; display: flex; gap: 12px; flex-wrap: wrap; }
.filters label { font-size: 12px; }
.filters input, .filters select { padding: 6px 10px; border: 1px solid #DDD; border-radius: 4px; font-size: 13px; }
table { width: 100%; border-collapse: collapse; background: white; font-size: 12px; }
th { background: #1F3864; color: white; padding: 8px 6px; text-align: right; position: sticky; top: 0; cursor: pointer; }
th:first-child, th:nth-child(2) { text-align: left; }
td { padding: 6px; border-bottom: 1px solid #EEE; text-align: right; }
td:first-child, td:nth-child(2) { text-align: left; }
td:first-child { font-weight: bold; color: #1F3864; }
.zebra:nth-child(even) { background: #FAFAFA; }
.pill { display: inline-block; padding: 2px 8px; border-radius: 10px; font-size: 10px; font-weight: bold; }
.pill-gov-d { background: #E3F2FD; color: #1565C0; }
.pill-gov-i { background: #F3E5F5; color: #6A1B9A; }
.pill-priv { background: #FFF3E0; color: #E65100; }
.warn { color: #C62828; font-weight: bold; }
.good { color: #2E7D32; font-weight: bold; }
</style></head><body>
<h1>Detalhamento de Faturas — Projeto PRODAM</h1>
<p style="color:#666">Breakdown por situação, prescrição e exigibilidade. Cutoff prescrição: """
        + f"{CUTOFF_MES:02d}/{CUTOFF_ANO}"
        + """ (5 anos - Art. 206 §5º I CC).</p>

<div class="kpis" id="kpis"></div>

<div class="filters">
  <label>Buscar devedor: <input id="fSearch" type="text" placeholder="SEDUC, SES..."></label>
  <label>Categoria: <select id="fCat"><option value="">Todas</option><option value="GOV_DIRETA">Gov Direta</option><option value="GOV_INDIRETA">Gov Indireta</option><option value="PRIVADA">Privada</option></select></label>
  <label><input type="checkbox" id="fSoPrescr"> Só com prescritas > 0</label>
  <label><input type="checkbox" id="fSoAtivo"> Só com exigíveis > 0</label>
</div>

<div style="overflow-x: auto; background: white; border-radius: 4px;">
<table id="tbl">
<thead><tr>
  <th data-sort="sigla">Devedor</th>
  <th data-sort="categoria">Cat</th>
  <th data-sort="total_qtd">Total Qtd</th>
  <th data-sort="total_val">Total Valor</th>
  <th data-sort="emitida_qtd">Emit Qtd</th>
  <th data-sort="emitida_val">Emit Val</th>
  <th data-sort="parcial_qtd">Parc Qtd</th>
  <th data-sort="parcial_val">Parc Val</th>
  <th data-sort="quitada_qtd">Quit Qtd</th>
  <th data-sort="cancelada_qtd">Canc Qtd</th>
  <th data-sort="prescritas_qtd">Prescr Qtd</th>
  <th data-sort="prescritas_val">Prescr Val</th>
  <th data-sort="nao_prescritas_qtd">Não-Pr Qtd</th>
  <th data-sort="nao_prescritas_val">Não-Pr Val</th>
  <th data-sort="exig_forte_qtd">Exig FORTE Qtd</th>
  <th data-sort="exig_forte_val">Exig FORTE Val</th>
</tr></thead>
<tbody id="tbody"></tbody>
</table>
</div>

<script>
const data = """
        + data_json
        + """;
const fmtBRL = v => "R$ " + v.toLocaleString("pt-BR", {minimumFractionDigits: 2, maximumFractionDigits: 2});
let sortKey = "total_val", sortAsc = false;

function filt() {
  const q = document.getElementById("fSearch").value.toUpperCase();
  const cat = document.getElementById("fCat").value;
  const soPr = document.getElementById("fSoPrescr").checked;
  const soAt = document.getElementById("fSoAtivo").checked;
  return data.filter(d =>
    (!q || d.sigla.toUpperCase().includes(q)) &&
    (!cat || d.categoria === cat) &&
    (!soPr || d.prescritas_qtd > 0) &&
    (!soAt || d.exig_forte_qtd > 0 || d.exig_media_qtd > 0)
  );
}

function render() {
  const rows = filt().sort((a, b) => {
    const va = a[sortKey], vb = b[sortKey];
    const cmp = typeof va === "number" ? va - vb : String(va).localeCompare(String(vb));
    return sortAsc ? cmp : -cmp;
  });

  // KPIs
  const tot = rows.reduce((s, d) => ({
    qtd: s.qtd + d.total_qtd, val: s.val + d.total_val,
    prescr_qtd: s.prescr_qtd + d.prescritas_qtd, prescr_val: s.prescr_val + d.prescritas_val,
    nao_prescr_qtd: s.nao_prescr_qtd + d.nao_prescritas_qtd, nao_prescr_val: s.nao_prescr_val + d.nao_prescritas_val,
    exig_forte_qtd: s.exig_forte_qtd + d.exig_forte_qtd, exig_forte_val: s.exig_forte_val + d.exig_forte_val,
    parc_qtd: s.parc_qtd + d.parcial_qtd, parc_val: s.parc_val + d.parcial_val,
    quit_qtd: s.quit_qtd + d.quitada_qtd,
  }), {qtd:0,val:0,prescr_qtd:0,prescr_val:0,nao_prescr_qtd:0,nao_prescr_val:0,exig_forte_qtd:0,exig_forte_val:0,parc_qtd:0,parc_val:0,quit_qtd:0});

  document.getElementById("kpis").innerHTML = `
    <div class="kpi"><div class="kpi-label">Devedores (filtro)</div><div class="kpi-val">${rows.length}</div></div>
    <div class="kpi"><div class="kpi-label">Total Faturas</div><div class="kpi-val">${tot.qtd}</div><div class="kpi-sub">${fmtBRL(tot.val)}</div></div>
    <div class="kpi"><div class="kpi-label">Não-Prescritas</div><div class="kpi-val good">${tot.nao_prescr_qtd}</div><div class="kpi-sub">${fmtBRL(tot.nao_prescr_val)}</div></div>
    <div class="kpi"><div class="kpi-label">Prescritas</div><div class="kpi-val warn">${tot.prescr_qtd}</div><div class="kpi-sub">${fmtBRL(tot.prescr_val)}</div></div>
    <div class="kpi"><div class="kpi-label">Exig. FORTE</div><div class="kpi-val good">${tot.exig_forte_qtd}</div><div class="kpi-sub">${fmtBRL(tot.exig_forte_val)}</div></div>
    <div class="kpi"><div class="kpi-label">Parcialmente Pagas</div><div class="kpi-val">${tot.parc_qtd}</div><div class="kpi-sub">${fmtBRL(tot.parc_val)}</div></div>
    <div class="kpi"><div class="kpi-label">Quitadas</div><div class="kpi-val">${tot.quit_qtd}</div></div>
  `;

  // Table
  const catPill = c => `<span class="pill pill-${c==='GOV_DIRETA'?'gov-d':c==='GOV_INDIRETA'?'gov-i':'priv'}">${c}</span>`;
  document.getElementById("tbody").innerHTML = rows.map(d => `
    <tr class="zebra">
      <td>${d.sigla}</td>
      <td>${catPill(d.categoria)}</td>
      <td>${d.total_qtd}</td>
      <td>${fmtBRL(d.total_val)}</td>
      <td>${d.emitida_qtd}</td>
      <td>${fmtBRL(d.emitida_val)}</td>
      <td>${d.parcial_qtd}</td>
      <td>${fmtBRL(d.parcial_val)}</td>
      <td>${d.quitada_qtd}</td>
      <td>${d.cancelada_qtd}</td>
      <td class="${d.prescritas_qtd>0?'warn':''}">${d.prescritas_qtd}</td>
      <td class="${d.prescritas_val>0?'warn':''}">${fmtBRL(d.prescritas_val)}</td>
      <td class="${d.nao_prescritas_qtd>0?'good':''}">${d.nao_prescritas_qtd}</td>
      <td class="${d.nao_prescritas_val>0?'good':''}">${fmtBRL(d.nao_prescritas_val)}</td>
      <td>${d.exig_forte_qtd}</td>
      <td>${fmtBRL(d.exig_forte_val)}</td>
    </tr>`).join("");
}

document.querySelectorAll("th[data-sort]").forEach(th => {
  th.onclick = () => {
    const k = th.dataset.sort;
    if (sortKey === k) sortAsc = !sortAsc;
    else { sortKey = k; sortAsc = false; }
    render();
  };
});
["fSearch", "fCat", "fSoPrescr", "fSoAtivo"].forEach(id => {
  document.getElementById(id).addEventListener("input", render);
  document.getElementById(id).addEventListener("change", render);
});
render();
</script>
</body></html>"""
    )
    (OUT / "DETALHAMENTO_FATURAS.html").write_text(html, encoding="utf-8")


def gerar_md(consolidado, sem_match, clientes_orfaos):
    tot = {
        "qtd": sum(c["total_qtd"] for c in consolidado),
        "val": sum(c["total_val"] for c in consolidado),
        "emit": (
            sum(c["emitida_qtd"] for c in consolidado),
            sum(c["emitida_val"] for c in consolidado),
        ),
        "parc": (
            sum(c["parcial_qtd"] for c in consolidado),
            sum(c["parcial_val"] for c in consolidado),
        ),
        "perd": (
            sum(c["perdas_qtd"] for c in consolidado),
            sum(c["perdas_val"] for c in consolidado),
        ),
        "quit": (
            sum(c["quitada_qtd"] for c in consolidado),
            sum(c["quitada_val"] for c in consolidado),
        ),
        "canc": (
            sum(c["cancelada_qtd"] for c in consolidado),
            sum(c["cancelada_val"] for c in consolidado),
        ),
        "presc": (
            sum(c["prescritas_qtd"] for c in consolidado),
            sum(c["prescritas_val"] for c in consolidado),
        ),
        "npresc": (
            sum(c["nao_prescritas_qtd"] for c in consolidado),
            sum(c["nao_prescritas_val"] for c in consolidado),
        ),
        "exf": (
            sum(c["exig_forte_qtd"] for c in consolidado),
            sum(c["exig_forte_val"] for c in consolidado),
        ),
        "exm": (
            sum(c["exig_media_qtd"] for c in consolidado),
            sum(c["exig_media_val"] for c in consolidado),
        ),
        "exfr": (
            sum(c["exig_fraca_qtd"] for c in consolidado),
            sum(c["exig_fraca_val"] for c in consolidado),
        ),
    }

    def brl(v):
        return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

    md = f"""# Detalhamento de Faturas — Projeto PRODAM

**Gerado em:** {HOJE:%d/%m/%Y}
**Cutoff prescrição:** {CUTOFF_MES:02d}/{CUTOFF_ANO} (5 anos - Art. 206 §5º I CC)

## Totais Globais (devedores do portfólio)

| Categoria | Qtd | Valor |
|-----------|-----|-------|
| **Total no DB (em devedores do portfólio)** | {tot['qtd']} | {brl(tot['val'])} |
| Emitida (em aberto) | {tot['emit'][0]} | {brl(tot['emit'][1])} |
| Parcialmente Paga | {tot['parc'][0]} | {brl(tot['parc'][1])} |
| EmPerdas | {tot['perd'][0]} | {brl(tot['perd'][1])} |
| Totalmente Paga | {tot['quit'][0]} | {brl(tot['quit'][1])} |
| Cancelada | {tot['canc'][0]} | {brl(tot['canc'][1])} |
| **Prescritas (≥5 anos)** | **{tot['presc'][0]}** | **{brl(tot['presc'][1])}** |
| **Não-prescritas** | **{tot['npresc'][0]}** | **{brl(tot['npresc'][1])}** |
| Exig. FORTE | {tot['exf'][0]} | {brl(tot['exf'][1])} |
| Exig. MÉDIA | {tot['exm'][0]} | {brl(tot['exm'][1])} |
| Exig. FRACA | {tot['exfr'][0]} | {brl(tot['exfr'][1])} |

## Top 10 Devedores por Valor Total

| # | Devedor | Cat | Qtd | Valor | Prescritas | Exig. FORTE |
|---|---------|-----|-----|-------|------------|-------------|
"""
    for i, c in enumerate(sorted(consolidado, key=lambda x: -x["total_val"])[:10], 1):
        md += f"| {i} | **{c['sigla']}** | {c['categoria']} | {c['total_qtd']} | {brl(c['total_val'])} | {c['prescritas_qtd']} / {brl(c['prescritas_val'])} | {c['exig_forte_qtd']} / {brl(c['exig_forte_val'])} |\n"

    md += f"\n## Siglas SSOT sem faturas no DB ({len(sem_match)})\n\n"
    if sem_match:
        md += "Estes devedores estão no SSOT (`profiles.json`) mas não têm faturas extraídas no `prodam.db`:\n\n"
        for s in sem_match:
            md += f"- {s}\n"

    md += f"\n## Clientes no DB fora do portfólio ({len(clientes_orfaos)})\n\n"
    md += "Estes clientes têm faturas no DB mas não são devedores do projeto (despachantes, auto-escolas, pessoas físicas, etc):\n\n"
    md += f"<details><summary>Lista completa ({len(clientes_orfaos)})</summary>\n\n"
    for c in clientes_orfaos:
        md += f"- `{c}`\n"
    md += "\n</details>\n"

    (OUT / "RESUMO_DETALHAMENTO.md").write_text(md, encoding="utf-8")


if __name__ == "__main__":
    main()
