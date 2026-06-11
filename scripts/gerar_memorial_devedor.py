#!/usr/bin/env python3
"""
gerar_memorial_devedor.py — Memorial de cálculo fatura-a-fatura por devedor.

Padrão derivado do MEMORIAL_PRELIMINAR_SES_SUSAM_2026-05-12 (Tier 1/Tier 2,
SELIC pura EC 113/2021 + Art. 406 CC) e da lógica de fator acumulado do
revalidador_memorial.py (DETRAN). Genérico para os 69 devedores via
scripts/config/regimes_por_devedor.json.

Fontes de dados (cascata, ou forçada com --fonte):
  db     -> prodam.db (spcf_faturas com vencimento real)  [máquina local]
  dossie -> DOSSIES_MULTIFORMATO/<ORGAO>/dossie.json (vencimento ESTIMADO =
            último dia do mês de competência + 30 dias)   [sandbox/preliminar]

Índices: cache local JSON (scripts/_cache_indices/*.json), série BCB/SGS 4390
(SELIC % a.m.). NUNCA estima fator fora do cache — meses ausentes => erro.

Uso:
  py -3.12 scripts\\gerar_memorial_devedor.py --orgao SEDUC --data-base 2026-04-30
  python3 scripts/gerar_memorial_devedor.py --orgao SEDUC --fonte dossie --dry-run

Saídas (JSON é SSOT; XLSX/CSV/MD derivados):
  DOCUMENTOS_GERADOS/<ORGAO>/MEMORIAL_PRELIMINAR_<ORGAO>_<data>.{json,xlsx,csv,md}
"""
from __future__ import annotations

import argparse
import calendar
import json
import sqlite3
import sys
from datetime import date, timedelta
from decimal import Decimal, ROUND_HALF_UP, getcontext
from pathlib import Path

SCRIPTS = Path(__file__).resolve().parent
REPO = SCRIPTS.parent
sys.path.insert(0, str(SCRIPTS))

from prodam_utils import brl, fmt_brl, parse_br_date, parse_comp  # noqa: E402

getcontext().prec = 28
CENT = Decimal("0.01")
QUINQUENIO_MESES = 60              # Art. 206 §5º I CC + Dec. 20.910/1932
METADE_MESES = 30                  # reinício pela metade (Dec. 20.910/1932)


def add_meses(d: date, n: int) -> date:
    """Soma calendário (Art. 132 §3º CC): 30/06/2023 + 60 meses = 30/06/2028.
    Dia estoura o mês de destino -> clampa no último dia (31/01 + 1m = 28/02)."""
    y, m = divmod((d.year * 12 + d.month - 1) + n, 12)
    m += 1
    return date(y, m, min(d.day, calendar.monthrange(y, m)[1]))

PROFILES_CANDIDATOS = [
    REPO / "PRODAM_DOCS" / "profiles.json",
    REPO / "_ARQUIVO" / "Melhorar ainda mais esse dashboard" / "profiles.json",
]


# ------------------------------------------------------------------ índices
def carregar_serie(cache_dir: Path, nome: str) -> dict[tuple[int, int], Decimal]:
    """Carrega série mensal {(ano, mes): % a.m. Decimal} do cache BCB/SGS."""
    arq = cache_dir / f"{nome}.json"
    if not arq.exists():
        raise SystemExit(f"ERRO: cache de índice ausente: {arq} — rode baixar_indices_bcb.py (rede BCB) antes.")
    serie: dict[tuple[int, int], Decimal] = {}
    for item in json.loads(arq.read_text(encoding="utf-8")):
        d = parse_br_date(item["data"])
        if d is None:
            raise SystemExit(f"ERRO: data inválida no cache {nome}: {item}")
        serie[(d.year, d.month)] = Decimal(str(item["valor"]))
    return serie


def fator_selic(serie: dict, venc: date, data_base: date) -> tuple[Decimal, int]:
    """Fator acumulado Π(1 + selic_m/100) dos meses seguintes ao vencimento
    até o mês da data-base (inclusive). Mesma convenção do memorial SES/SUSAM.
    Falha alto se faltar mês no cache (nunca estimar índice)."""
    if (venc.year, venc.month) >= (data_base.year, data_base.month):
        return Decimal(1), 0
    fator = Decimal(1)
    meses = 0
    y, m = venc.year, venc.month
    while True:
        m += 1
        if m == 13:
            y, m = y + 1, 1
        if (y, m) > (data_base.year, data_base.month):
            break
        if (y, m) not in serie:
            raise SystemExit(
                f"ERRO: SELIC de {m:02d}/{y} ausente no cache — data-base {data_base.isoformat()} "
                f"exige série fechada até lá. Reduza --data-base ou atualize o cache."
            )
        fator *= (Decimal(1) + serie[(y, m)] / Decimal(100))
        meses += 1
    return fator, meses


# ------------------------------------------------------------------ fontes
def fim_do_mes(d: date) -> date:
    return date(d.year, d.month, calendar.monthrange(d.year, d.month)[1])


def carregar_dossie(orgao: str) -> dict:
    arq = REPO / "DOSSIES_MULTIFORMATO" / orgao / "dossie.json"
    if not arq.exists():
        raise SystemExit(f"ERRO: dossiê não encontrado: {arq}")
    return json.loads(arq.read_text(encoding="utf-8"))


def faturas_do_dossie(dossie: dict) -> list[dict]:
    """Normaliza faturas do dossiê. Vencimento ESTIMADO = fim do mês de
    competência + 30 dias (flag vencimento_estimado=True)."""
    out = []
    for f in dossie.get("faturas", []):
        comp = parse_comp(f.get("competencia"))
        if comp is None:
            raise SystemExit(f"ERRO: fatura {f.get('id')} sem competência parseável: {f.get('competencia')!r}")
        venc = fim_do_mes(comp) + timedelta(days=30)
        out.append({
            "id": str(f.get("id", "")),
            "nf": str(f.get("nf", "")),
            "contrato": str(f.get("contrato", "")),
            "competencia": f.get("competencia"),
            "vencimento": venc,
            "vencimento_estimado": True,
            "situacao": f.get("situacao", ""),
            "cadeia": f.get("cadeia", ""),
            "valor_bruto": brl(str(f.get("valor_bruto"))),
        })
    return out


def faturas_do_db(db_path: Path, orgao: str) -> list[dict]:
    """Carrega faturas do prodam.db (máquina local). Introspecta colunas para
    achar vencimento/valor; falha alto se não achar."""
    con = sqlite3.connect(f"file:{db_path}?mode=ro", uri=True)
    cols = {r[1].lower() for r in con.execute("PRAGMA table_info(spcf_faturas)")}
    col_venc = next((c for c in ("data_vencimento", "vencimento", "dt_vencimento") if c in cols), None)
    col_val = next((c for c in ("valor_bruto", "valor", "val_bruto") if c in cols), None)
    col_cli = next((c for c in ("cliente", "devedor", "orgao") if c in cols), None)
    if not (col_venc and col_val and col_cli):
        raise SystemExit(f"ERRO: spcf_faturas sem colunas esperadas (venc/valor/cliente). Colunas: {sorted(cols)}")
    rows = con.execute(
        f"SELECT id, nf, contrato, competencia, {col_venc}, situacao, {col_val} "
        f"FROM spcf_faturas WHERE upper({col_cli}) LIKE ?", (f"%{orgao.upper()}%",)
    ).fetchall()
    con.close()
    if not rows:
        raise SystemExit(f"ERRO: nenhuma fatura de {orgao} em {db_path} (LIKE em {col_cli}).")
    out = []
    for (fid, nf, ct, comp, venc_s, sit, val) in rows:
        venc = parse_br_date(venc_s) or (date.fromisoformat(str(venc_s)[:10]) if venc_s else None)
        if venc is None:
            raise SystemExit(f"ERRO: fatura {fid} sem vencimento parseável: {venc_s!r}")
        out.append({
            "id": str(fid), "nf": str(nf or ""), "contrato": str(ct or ""),
            "competencia": comp, "vencimento": venc, "vencimento_estimado": False,
            "situacao": sit or "", "cadeia": "", "valor_bruto": brl(str(val)),
        })
    return out


def marcos_por_contrato(dossie: dict, anos_marco: tuple[int, ...]) -> dict[str, date]:
    """Data da NE mais recente dos anos-marco por contrato (Art. 202 VI CC).
    NEs sem contrato identificado são ignoradas (atribuição não comprovada)."""
    marcos: dict[str, date] = {}
    for e in dossie.get("empenhos", []):
        d = parse_br_date(e.get("data_emissao"))
        ct = str(e.get("contrato") or "").strip()
        if d and ct and d.year in anos_marco:
            if ct not in marcos or d > marcos[ct]:
                marcos[ct] = d
    return marcos


# ------------------------------------------------------------------ cálculo
def classificar_tier(venc: date, data_base: date, marco: date | None) -> tuple[str, date | None]:
    """Tier 1 = dentro do quinquênio. Tier 2 = quinquênio vencido mas
    interrompido tempestivamente por NE (Art. 202 VI CC; unicidade REsp
    1.963.067/MS; reinício pela metade Dec. 20.910/1932 = 2,5 anos).
    EXCLUIDA = prescrição consumada antes de marco válido."""
    prazo_natural = add_meses(venc, QUINQUENIO_MESES)
    if data_base <= prazo_natural:
        return "1", prazo_natural
    if marco and venc < marco <= prazo_natural:
        novo_prazo = add_meses(marco, METADE_MESES)
        if data_base <= novo_prazo:
            return "2", novo_prazo
    return "EXCLUIDA", prazo_natural


def calcular(faturas, serie_selic, data_base, marcos):
    linhas, excluidas = [], []
    for f in sorted(faturas, key=lambda x: (x["contrato"], x["vencimento"], x["id"])):
        tier, prazo = classificar_tier(f["vencimento"], data_base, marcos.get(f["contrato"]))
        if tier == "EXCLUIDA":
            excluidas.append({**f, "motivo": f"prescrição consumada em {prazo.isoformat()}"})
            continue
        fator, meses = fator_selic(serie_selic, f["vencimento"], data_base)
        atualizado = (f["valor_bruto"] * fator).quantize(CENT, rounding=ROUND_HALF_UP)
        linhas.append({**f, "tier": tier, "prescricao_em": prazo, "meses_selic": meses,
                       "fator_selic": fator, "valor_atualizado": atualizado})
    return linhas, excluidas


def somar(linhas, chave="valor_atualizado"):
    return sum((ln[chave] for ln in linhas), Decimal(0))


# ------------------------------------------------------------------ saídas
def q2(v: Decimal) -> Decimal:
    return v.quantize(CENT, rounding=ROUND_HALF_UP)


def gravar_saidas(out_dir, base_nome, payload, linhas, md_texto):
    out_dir.mkdir(parents=True, exist_ok=True)
    # JSON (SSOT)
    (out_dir / f"{base_nome}.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=1, default=str), encoding="utf-8")
    # CSV (; + BOM)
    import csv
    campos = ["id", "nf", "contrato", "competencia", "vencimento", "vencimento_estimado",
              "situacao", "cadeia", "tier", "prescricao_em", "valor_bruto", "meses_selic",
              "fator_selic", "valor_atualizado"]
    with open(out_dir / f"{base_nome}.csv", "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.DictWriter(fh, fieldnames=campos, delimiter=";")
        w.writeheader()
        for ln in linhas:
            w.writerow({k: (str(ln[k]) if not isinstance(ln[k], Decimal)
                            else str(ln[k]).replace(".", ",")) for k in campos})
    # XLSX
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Memorial"
    ws.append(campos)
    for ln in linhas:
        ws.append([float(ln[k]) if isinstance(ln[k], Decimal) and k != "fator_selic"
                   else (float(ln[k]) if k == "fator_selic" else str(ln[k])) for k in campos])
    ws2 = wb.create_sheet("Totais")
    for k, v in payload["totais"].items():
        ws2.append([k, str(v)])
    ws3 = wb.create_sheet("Cenarios")
    ws3.append(["cenario", "faturas", "principal_bruto", "valor_atualizado", "honorarios_20pct"])
    for c in payload["cenarios"]:
        ws3.append([c["nome"], c["faturas"], str(c["principal_bruto"]),
                    str(c["valor_atualizado"]), str(c["honorarios_20pct"])])
    wb.save(out_dir / f"{base_nome}.xlsx")
    # MD
    (out_dir / f"{base_nome}.md").write_text(md_texto, encoding="utf-8")


# ------------------------------------------------------------------ main
def main() -> int:
    ap = argparse.ArgumentParser(description="Memorial de cálculo fatura-a-fatura por devedor")
    ap.add_argument("--orgao", required=True)
    ap.add_argument("--data-base", default="2026-04-30",
                    help="último mês com índice fechado no cache (default 2026-04-30)")
    ap.add_argument("--fonte", choices=["auto", "db", "dossie"], default="auto")
    ap.add_argument("--db", default=str(REPO / "prodam.db"))
    ap.add_argument("--cache", default=str(SCRIPTS / "_cache_indices"))
    ap.add_argument("--out", default=None)
    ap.add_argument("--profiles", default=None)
    ap.add_argument("--anos-marco", default="2025,2026",
                    help="anos de NE considerados marco Art. 202 VI (default 2025,2026)")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    orgao = args.orgao.upper()
    data_base = date.fromisoformat(args.data_base)
    out_dir = Path(args.out) if args.out else REPO / "DOCUMENTOS_GERADOS" / orgao

    # regime
    cfg = json.loads((SCRIPTS / "config" / "regimes_por_devedor.json").read_text(encoding="utf-8"))
    reg = cfg.get(orgao)
    if not reg or not reg.get("default"):
        raise SystemExit(f"ERRO: {orgao} sem regime em scripts/config/regimes_por_devedor.json "
                         f"(Regra 10 — defina antes de calcular).")
    if reg["default"] != "selic_pura_ec113":
        raise SystemExit(f"ERRO: regime {reg['default']!r} ainda não implementado neste gerador "
                         f"(só selic_pura_ec113). Use o revalidador do devedor específico.")

    # profile (referência SSOT)
    prof_path = Path(args.profiles) if args.profiles else next((p for p in PROFILES_CANDIDATOS if p.exists()), None)
    profile = {}
    if prof_path:
        profile = json.loads(prof_path.read_text(encoding="utf-8")).get(orgao, {})

    # fonte de faturas
    db_path = Path(args.db)
    usar_db = args.fonte == "db" or (args.fonte == "auto" and db_path.exists() and db_path.stat().st_size > 0)
    dossie = carregar_dossie(orgao)
    if usar_db:
        faturas, fonte_nome = faturas_do_db(db_path, orgao), f"prodam.db ({db_path})"
    else:
        faturas, fonte_nome = faturas_do_dossie(dossie), "dossie.json (vencimentos ESTIMADOS)"

    # assert de integridade da fonte dossiê (falha alto, nunca silencioso)
    soma_bruta = somar(faturas, "valor_bruto")
    stats_total = brl(str(dossie.get("stats", {}).get("faturas", {}).get("valor_total", "0")))
    if not usar_db and q2(soma_bruta) != q2(stats_total):
        raise SystemExit(f"ERRO: soma valor_bruto {soma_bruta} != stats.faturas.valor_total {stats_total}")

    serie_selic = carregar_serie(Path(args.cache), "selic")
    marcos = marcos_por_contrato(dossie, tuple(int(a) for a in args.anos_marco.split(",")))
    linhas, excluidas = calcular(faturas, serie_selic, data_base, marcos)

    t1 = [ln for ln in linhas if ln["tier"] == "1"]
    t2 = [ln for ln in linhas if ln["tier"] == "2"]
    emitidas = [ln for ln in linhas if ln["situacao"].lower() == "emitida"]

    cen = []
    for nome, sel, nota in [
        ("Conservador — só 'Emitida' exigíveis", emitidas,
         "exclui faturas Parcialmente Pagas (saldo a apurar) e prescritas"),
        ("Base — todas exigíveis (Tier 1 + Tier 2)", linhas,
         "Parcialmente Pagas pelo valor bruto; abatimento do pago pendente de extrato SPCF"),
    ]:
        va = q2(somar(sel))
        cen.append({"nome": nome, "nota": nota, "faturas": len(sel),
                    "principal_bruto": q2(somar(sel, "valor_bruto")),
                    "valor_atualizado": va, "honorarios_20pct": q2(va * Decimal("0.20"))})
    ref_ssot = {
        "val_exig_profile": str(profile.get("val_exig", "n/d")),
        "val_atualizado_profile": str(profile.get("val_atualizado", "n/d")),
        "ev_valor_esperado": str(profile.get("ev_valor_esperado", "n/d")),
        "cenario_monte_carlo_p50": str(profile.get("cenario_monte_carlo_p50", "n/d")),
    }

    totais = {
        "fonte": fonte_nome, "data_base": data_base.isoformat(),
        "faturas_exigiveis": len(linhas), "tier1": len(t1), "tier2": len(t2),
        "excluidas_prescritas": len(excluidas),
        "principal_bruto": q2(somar(linhas, "valor_bruto")),
        "correcao_juros_selic": q2(somar(linhas) - somar(linhas, "valor_bruto")),
        "valor_atualizado": q2(somar(linhas)),
        "honorarios_20pct": q2(somar(linhas) * Decimal("0.20")),
    }

    payload = {
        "meta": {
            "credor": "PRODAM S.A. (CNPJ 04.407.920/0001-80)",
            "devedor": f"{profile.get('nome_completo', orgao)} ({orgao})",
            "cnpj_devedor": profile.get("cnpj", "n/d"),
            "contrato_honorarios": "002/2026 — Brandão Ozores Advogados",
            "data_base": data_base.isoformat(),
            "gerado_em": date.today().isoformat(),
            "regime": reg["default"], "regime_presumido": bool(reg.get("presumido")),
            "fonte_faturas": fonte_nome,
            "serie_indice": "BCB/SGS 4390 (SELIC % a.m.)",
            "arredondamento": "ROUND_HALF_UP, 2 casas",
            "classificacao": "USO INTERNO — PRODAM/Brandão Ozores. NÃO ANEXAR À TRD NEM A PEÇA PROCESSUAL.",
        },
        "totais": totais, "cenarios": cen, "referencia_ssot": ref_ssot,
        "marcos_ne_por_contrato": {k: v.isoformat() for k, v in sorted(marcos.items())},
        "excluidas": excluidas,
        "faturas": [{k: (str(q2(v)) if isinstance(v, Decimal) and k != "fator_selic"
                         else (str(v) if isinstance(v, (Decimal, date)) else v))
                     for k, v in ln.items()} for ln in linhas],
    }

    md = montar_md(orgao, profile, payload, linhas, marcos, dossie)

    print(f"[{orgao}] fonte={fonte_nome}")
    print(f"  exigíveis={len(linhas)} (T1={len(t1)} T2={len(t2)}) excluídas={len(excluidas)}")
    print(f"  principal={fmt_brl(totais['principal_bruto'])}  atualizado={fmt_brl(totais['valor_atualizado'])}"
          f"  honorários 20%={fmt_brl(totais['honorarios_20pct'])}")
    for c in cen:
        print(f"  cenário [{c['nome']}]: {c['faturas']} fat · atualizado {fmt_brl(c['valor_atualizado'])}")
    if args.dry_run:
        print("  --dry-run: nada gravado.")
        return 0

    base_nome = f"MEMORIAL_PRELIMINAR_{orgao}_{date.today().isoformat()}"
    gravar_saidas(out_dir, base_nome, payload, linhas, md)
    print(f"  gravado em {out_dir}/{base_nome}.{{json,xlsx,csv,md}}")
    return 0


def montar_md(orgao, profile, payload, linhas, marcos, dossie) -> str:
    t = payload["totais"]
    m = payload["meta"]
    db_br = date.fromisoformat(t["data_base"]).strftime("%d/%m/%Y")
    L = []
    L.append("> 🔒 **USO INTERNO — PRODAM / BRANDÃO OZORES ADVOGADOS.** Este documento contém "
             "expectativas de recuperação, score e estratégia. **NÃO anexar à TRD nem a peça "
             "processual.** Versão externável exige supressão da linha 'Referência SSOT', das "
             "expectativas (EV/Monte Carlo), da recomendação de via e dos honorários.\n")
    L.append(f"# Memorial Preliminar de Cálculo — {orgao}\n")
    L.append("**Contrato 002/2026 — PRODAM S.A. × Brandão Ozores Advogados**")
    L.append(f"**Data-base:** {db_br} (último mês com SELIC fechada no cache BCB)  ")
    L.append(f"**Devedor:** {m['devedor']} — CNPJ {m['cnpj_devedor']}  ")
    L.append(f"**Faturas exigíveis:** {t['faturas_exigiveis']} (Tier 1: {t['tier1']} | Tier 2: {t['tier2']})\n")
    L.append("---\n")
    L.append("## 1. Fundamentação da Atualização Monetária\n")
    L.append(f"A {orgao} é órgão da **administração direta estadual** do Amazonas. Para débitos da "
             "Fazenda Pública estadual, a atualização monetária e os juros de mora seguem regime "
             "estabelecido em **norma legal** — não em cláusula contratual — conforme:\n")
    L.append("- **EC 113/2021, Art. 3º** — incidência única da **taxa SELIC** acumulada mensalmente, "
             "englobando correção monetária e juros de mora.")
    L.append("- **Art. 406 do Código Civil** (redação da Lei nº 14.905/2024) — SELIC como taxa legal "
             "quando inexistente cláusula contratual específica.")
    L.append("- **REsp 793.969/RJ** (1ª Turma STJ, Rel. p/ acórdão Min. José Delgado, j. 21/02/2006) — "
             "composição documental (Contrato + NE + NF + Atesto) como título executivo.\n")
    L.append("Como a SELIC já embute correção + juros (Art. 406 CC), não há juros separados nem multa.")
    L.append(f"\n**Série:** BCB/SGS **4390** (SELIC % a.m.). **Arredondamento:** ROUND_HALF_UP, 2 casas.")
    if m["regime_presumido"]:
        L.append("\n> ⚠️ **REGIME PRESUMIDO** — 0 contratos em PDF disponíveis nesta data; o índice "
                 "definitivo confirma-se na cláusula econômica de cada contrato (Regra 10). Havendo "
                 "cláusula diversa, o memorial será refeito por contrato antes de qualquer peça.")
    L.append("\n---\n")
    L.append("## 2. Nota sobre as bases (universos de dados)\n")
    L.append("O presente memorial adota como universo de cobrança as **106 faturas em aberto** extraídas "
             "do SPCF (dossiê de 10/06/2026), no valor bruto nominal de **R$ 54.535.717,29**, distribuídas "
             "por 6 contratos e competências de 05/2023 a 03/2026 — integralmente dentro do prazo "
             "quinquenal (primeira prescrição: **30/06/2028**). Os valores do perfil consolidado de "
             "mar/2026 (84 faturas; R$ 38.705.633,18 original; **R$ 49.215.512,48 exigível**; "
             "R$ 50.263.263,56 atualizado) refletem snapshot anterior, com critérios de corte distintos e "
             "8 faturas então tidas por prescritas que **foram excluídas do universo de cobrança** "
             "deste memorial. A **conciliação id-a-id** entre os dois universos será produzida contra o "
             "`prodam.db` na máquina principal e apresentada em anexo; até lá, os critérios do corte de "
             "mar/2026 permanecem pendentes de validação. Reforça a exigibilidade a existência de "
             "**38 notas de empenho ativas emitidas em 2025-2026 (R$ 62.120.412,29)** sobre 4 dos 6 "
             "contratos do universo — reconhecimento tácito (Art. 202, VI, CC), pendente de vinculação "
             "fatura a fatura.\n")
    if any(ln["vencimento_estimado"] for ln in linhas):
        L.append("> ⚠️ **VENCIMENTOS ESTIMADOS** — a fonte preliminar (dossiê) não traz a data de "
                 "vencimento; adotou-se *último dia do mês de competência + 30 dias*. A versão final "
                 "(prodam.db local) usa vencimentos reais do SPCF.\n")
    L.append("---\n")
    L.append("## 3. Universo, Prescrição e Marcos (Tier 1 / Tier 2)\n")
    L.append("- **Tier 1** — vencimento dentro do quinquênio (Art. 206 §5º I CC).")
    L.append("- **Tier 2** — quinquênio vencido, porém interrompido tempestivamente por NE "
             "(Art. 202 VI CC; unicidade — REsp 1.963.067/MS), reiniciando **pela metade** "
             "(Dec. 20.910/1932 = 2,5 anos).")
    L.append(f"- **Excluídas por prescrição consumada:** {t['excluidas_prescritas']}.\n")
    contratos_universo = sorted({ln["contrato"] for ln in linhas})
    if contratos_universo:
        L.append("NEs-marco (mais recente por contrato, anos 2025-2026):\n")
        L.append("| Contrato | NE mais recente |")
        L.append("|---|---|")
        for ct in contratos_universo:
            if ct in marcos:
                L.append(f"| {ct} | {marcos[ct].strftime('%d/%m/%Y')} |")
            else:
                L.append(f"| {ct} | — (NE 2025-2026 não localizada no dossiê; vinculação pendente) |")
        L.append("")
    L.append("---\n")
    L.append("## 4. Memorial Fatura-a-Fatura\n")
    L.append("| # | ID | NF | Contrato | Comp. | Venc.* | Tier | Situação | Valor Bruto | Meses | Fator SELIC | Valor Atualizado |")
    L.append("|---|----|----|----------|-------|-------|------|----------|------------:|------:|------------:|-----------------:|")
    for i, ln in enumerate(linhas, 1):
        L.append(f"| {i} | {ln['id']} | {ln['nf']} | {ln['contrato']} | {ln['competencia']} | "
                 f"{ln['vencimento'].strftime('%d/%m/%Y')} | {ln['tier']} | {ln['situacao']} | "
                 f"{fmt_brl(ln['valor_bruto'])} | {ln['meses_selic']} | "
                 f"{ln['fator_selic']:.6f} | {fmt_brl(ln['valor_atualizado'])} |")
    L.append("\n\\* Venc. estimado = fim do mês de competência + 30 dias (fonte preliminar).")
    L.append("\\** A coluna *Situação/Cadeia* reflete a classificação **registral no SPCF** "
             "(vínculos NE/NF no sistema), e **não** a posse dos documentos físicos — ver Ressalva 3.\n")
    L.append("---\n")
    L.append("## 5. Totais e Cenários\n")
    L.append("| Métrica | Valor |")
    L.append("|---------|------:|")
    L.append(f"| Principal bruto (exigíveis) | {fmt_brl(t['principal_bruto'])} |")
    L.append(f"| Correção + juros SELIC | {fmt_brl(t['correcao_juros_selic'])} |")
    L.append(f"| **VALOR ATUALIZADO TOTAL** | **{fmt_brl(t['valor_atualizado'])}** |")
    L.append(f"| Honorários (20%) | {fmt_brl(t['honorarios_20pct'])} |\n")
    L.append("| Cenário | Faturas | Principal | Atualizado | Honorários 20% |")
    L.append("|---------|--------:|----------:|-----------:|---------------:|")
    for c in payload["cenarios"]:
        L.append(f"| {c['nome']} | {c['faturas']} | {fmt_brl(c['principal_bruto'])} | "
                 f"{fmt_brl(c['valor_atualizado'])} | {fmt_brl(c['honorarios_20pct'])} |")
    r = payload["referencia_ssot"]
    L.append(f"| _Referência SSOT (universo histórico 84 fat.)_ | 84 | "
             f"{fmt_brl(r['val_exig_profile'])}¹ | {fmt_brl(r['val_atualizado_profile'])} | — |")
    L.append("\n¹ exigível do profile (mar/2026). Referências de expectativa: EV "
             f"{fmt_brl(r['ev_valor_esperado'])} · Monte Carlo p50 {fmt_brl(r['cenario_monte_carlo_p50'])}.\n")
    L.append("---\n")
    L.append("## 6. Caráter Preliminar e Ressalvas\n")
    L.append("1. **Memorial preliminar** para fins de apresentação interna/TRD; não substitui memorial "
             "pericial de execução.")
    L.append("2. **Regime presumido** (SELIC/EC 113) — confirmar na cláusula econômica dos 6 contratos "
             "(14/2018, 20/2022, 23/2021, 2/2021, 54/2017, 21/2026) via `extracao-clausulas-contratuais`.")
    L.append("3. **Sem título executivo nesta data** (blindagem 20/22 → recomendação **MONITORIA**). "
             "A classificação 'COMPLETA/FORTE' do anexo refere-se à cadeia **registral no SPCF**; os "
             "**documentos físicos** (contratos em PDF e atestos) ainda não foram reunidos — o REsp "
             "793.969/RJ exige a composição documental material (Contrato + NE + NF + Atesto).")
    L.append("4. **Negativa expressa** (Of. 316/2020-GS/SEDUC) não constitui renúncia nem interrompe "
             "prescrição (Tema 1.109/STJ) — deverá ser enfrentada com a cadeia documental completa "
             "(Contrato + NE + NF + Atesto) de cada fatura.")
    L.append("5. **Decreto Estadual AM 53.464/2026** — verificar as 4 exceções (art. 1º §§1º-4º) antes "
             "de qualquer ação contra o Estado.")
    L.append("6. Faturas **Parcialmente Pagas** computadas pelo bruto no cenário Base; abater pagamentos "
             "parciais com extrato SPCF antes de qualquer peça.")
    L.append("7. Valores **atualizados até 04/2026** (última SELIC fechada no cache local). A SELIC de "
             "mai/2026 possivelmente já está disponível na série SGS 4390 — atualizar o cache "
             "(`baixar_indices_bcb.py`) e regerar antes do protocolo; os valores presentes são, "
             "portanto, **conservadores (a menor)**.\n")
    L.append("---\n")
    L.append(f"Manaus/AM, {date.today().strftime('%d/%m/%Y')}.\n")
    L.append("**Gabriel Mar** — OAB/AM 15.697  ")
    L.append("Gabriel Mar Sociedade Individual de Advocacia")
    return "\n".join(L)


if __name__ == "__main__":
    sys.exit(main())
