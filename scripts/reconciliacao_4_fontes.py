#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
reconciliacao_4_fontes.py — Reconciliação de valores por devedor cruzando 4 fontes

Fontes:
  1. profiles.json (SSOT) — 67 devedores, valores agregados
  2. prodam.db/spcf_faturas — faturas individuais com valor_bruto
  3. SPCF_EXTRACAO/por_devedor/ — JSONs crus do sistema SPCF
  4. PRODAM_DOCS/*_DOSSIE/INVENTARIO.xlsx — dossiês documentais

Gera:
  - reconciliacao_4_fontes_RELATORIO.md (relatório legível)
  - reconciliacao_4_fontes_DADOS.json (dados estruturados)
  - reconciliacao_4_fontes_RESUMO.xlsx (planilha consolidada)

Projeto PRODAM — Contrato 002/2026
"""

import json
import os
import sys
import sqlite3
import shutil
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from datetime import datetime
from pathlib import Path
from typing import Optional

# Helpers compartilhados — norm() resolve POLÍCIA CIVIL vs POLICIA CIVIL
BASE_DIR = Path(__file__).parent.parent
sys.path.insert(0, str(BASE_DIR))
sys.path.insert(0, str(BASE_DIR / "scripts"))
from prodam_utils import norm, norm_variants

# ============================================================
# CONFIGURAÇÃO
# ============================================================

PROFILES_PATH = BASE_DIR / "PRODAM_DOCS" / "profiles.json"
DB_PATH = BASE_DIR / "prodam.db"
SPCF_POR_DEVEDOR = BASE_DIR / "SPCF_EXTRACAO" / "por_devedor"
DOSSIE_DIR = BASE_DIR / "PRODAM_DOCS"
OUTPUT_DIR = BASE_DIR

# ============================================================
# NORMALIZAÇÃO DE SIGLAS
# ============================================================

# Mapa: sigla profiles.json → possíveis siglas nas outras fontes
SIGLA_MAP_DB_FATURAS = {
    "SES/SUSAM": ["SES"],
    "POLÍCIA CIVIL": ["POLICIA CIVIL"],
    "FAAR/SEDEL": ["FAAR - EXTINTA LEI 6.225 INCP PELA SEDEL", "SEDEL", "FAAR"],
    "FUAM/FUHAM": ["FUHAM", "FUAM"],
    "FMT": ["FMT-HVD", "FMT"],
    "BANCO MASTER": ["BANCO MASTER S/A", "BANCO MASTER"],
    "BANCO BMG": ["BANCO BMG"],
    "BANCO DAYCOVAL": ["BANCO DAYCOVAL"],
    "BANCO SAFRA": ["BANCO SAFRA"],
    "BANCO SICOOB": ["BANCO SICOOB"],
    "ASSOC. DE GESTÃO INOVAÇÃO E RES. EM SAÚDE": [
        "ASSOCIAÇÃO DE GESTÃO INOVAÇÃO E RESULTADOS EM SAUDE",
        "ASSOC. DE GESTÃO INOVAÇÃO E RES. EM SAÚDE",
    ],
    "B23 TECNOLOGIA": ["B23 TECNOLOGIA E PAGAMENTOS LTDA", "B23 TECNOLOGIA"],
    "PROVER": ["PROVER PROMOÇÃO DE VENDAS LTDA.", "PROVER"],
    "IKM DE": ["IKM DE LIMA-ME", "IKM DE"],
    "PSA TECHNOLOGY": ["PSA TECHNOLOGY"],
    "EASYTECH": ["EASYTECH"],
    "FENIXSOFT": ["FENIXSOFT", "FENIXSOFT2"],
    "ANOREG": ["ANOREG/AM", "ANOREG"],
    "COSAMA": ["COSAMA"],
    "FMPES": ["FMPES"],
    "FJJA": ["FJJA"],
    "SNPH": ["SNPH"],
    "FHAJ": ["FHAJ"],
    "SUL AMERICA": ["SUL AMERICA"],
    "ODONTOMED": ["ODONTOMED"],
    "ITRANSITO": ["ITRANSITO"],
    "SEJEL": ["SEJEL - EXTINTA LEI 122/2019", "SEJEL"],
    "SETRAB": ["SETRAB EXTINTA - LEI D. 122/19 - INC PELA SEDECTI", "SETRAB"],
    "SECT": ["SECT", "SECTI-EXTINTA"],
    "SEMIG": ["SEMIG"],
    "UGPI": ["UGPE - UNIDADE GESTORA DE PROJETOS ESPECIAIS", "UGPADEAM", "UGPI"],
    "SEAD": ["SEAD"],
    "CGE": ["CGE"],
    "CBMAM": ["CBMAM"],
    "CETAM": ["CETAM"],
    "CASA CIVIL": ["CASA CIVIL"],
    "CASA MILITAR": ["CASA MILITAR"],
    "DPE": ["DPE"],
    "PGE": ["PGE"],
    "PMAM": ["PMAM"],
    "SSP": ["SSP"],
    "SUHAB": ["SUHAB"],
    "DETRAN": ["DETRAN"],
    "SEDUC": ["SEDUC"],
    "SEFAZ": ["SEFAZ"],
    "SEINFRA": ["SEINFRA"],
    "SEJUSC": ["SEJUSC"],
    "SEDECTI": ["SEDECTI"],
    "SEAS": ["SEAS"],
    "SEC": ["SEC"],
    "SEMA": ["SEMA"],
    "SEPROR": ["SEPROR"],
    "IDAM": ["IDAM"],
    "IPAAM": ["IPAAM"],
    "FCECON": ["FCECON"],
    "FHEMOAM": ["FHEMOAM"],
    "FUNTEA": ["FUNTEA"],
    "FVS": ["FVS"],
    "AMAZONPREV": ["AMAZONPREV"],
    "CPA": ["CPA"],
    "CAIXA": ["CAIXA"],
    "BMC": ["BMC"],
    "ADS": ["ADS"],
    "ADAF": ["ADAF"],
    "ARSEPAM": ["ARSEPAM"],
    "AADESAM": ["AADESAM"],
}

SIGLA_MAP_SPCF_DIR = {
    "SES/SUSAM": "SES_SUSAM",
    "POLÍCIA CIVIL": "POLÍCIA CIVIL",
    "FAAR/SEDEL": "FAAR_SEDEL",
    "FUAM/FUHAM": "FUAM_FUHAM",
    "ASSOC. DE GESTÃO INOVAÇÃO E RES. EM SAÚDE": None,  # sem pasta dedicada
    "FMPES": None,
    "FJJA": None,
}

SIGLA_MAP_DOSSIE = {
    "SES/SUSAM": ["SES"],
    "POLÍCIA CIVIL": ["POLICIA_CIVIL"],
    "FAAR/SEDEL": ["FAAR", "SEDEL"],
    "FUAM/FUHAM": ["FUAM", "FUHAM"],
    "BANCO MASTER": ["BANCO_MASTER", "MASTER"],
    "BANCO BMG": ["BANCO_BMG"],
    "BANCO DAYCOVAL": [],  # pode não ter dossiê
    "BANCO SAFRA": ["BANCO_SAFRA"],
    "BANCO SICOOB": ["BANCO_SICOOB"],
    "ASSOC. DE GESTÃO INOVAÇÃO E RES. EM SAÚDE": ["ASSOC_SAUDE"],
    "B23 TECNOLOGIA": ["B23"],
    "IKM DE": ["IKM"],
    "PSA TECHNOLOGY": ["PSA"],
    "SUL AMERICA": ["SUL_AMERICA"],
}


def normalizar_brl(texto: str) -> Optional[Decimal]:
    """Converte string BRL para Decimal."""
    if not texto or texto.strip() in ("", "0", "0,00", "0.00"):
        return Decimal("0")
    try:
        s = str(texto).strip()
        s = s.replace("R$", "").replace("\xa0", "").strip()
        s = s.rstrip(".,")
        # Detectar formato BR (vírgula decimal)
        if "," in s and ("." in s or len(s.split(",")[-1]) <= 2):
            s = s.replace(".", "").replace(",", ".")
        elif "," in s:
            s = s.replace(",", ".")
        return Decimal(s).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError):
        return None


def fmt_brl(valor: Decimal) -> str:
    """Formata Decimal como R$ X.XXX,XX"""
    if valor is None:
        return "N/D"
    s = f"{valor:,.2f}"
    s = s.replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R$ {s}"


def pct_diff(a: Decimal, b: Decimal) -> Optional[float]:
    """Calcula % de diferença entre dois valores."""
    if a is None or b is None:
        return None
    if b == 0:
        return None if a == 0 else 100.0
    return float(abs(a - b) / b * 100)


# ============================================================
# FONTE 1: PROFILES.JSON
# ============================================================

def carregar_profiles() -> dict:
    """Carrega profiles.json (SSOT)."""
    with open(PROFILES_PATH, "r", encoding="utf-8") as f:
        data = json.load(f)
    # Remover _metadata
    data.pop("_metadata", None)
    result = {}
    for sigla, p in data.items():
        result[sigla] = {
            "val_exig": Decimal(str(p.get("val_exig", "0") or "0")),
            "val_orig": Decimal(str(p.get("val_orig", "0") or "0")),
            "val_atualizado": Decimal(str(p.get("val_atualizado", "0") or "0")),
            "faturas_total": int(p.get("faturas_total", 0) or 0),
            "faturas_exigiveis": int(p.get("faturas_exigiveis", 0) or 0),
            "faturas_prescritas": int(p.get("faturas_prescritas", 0) or 0),
            "forca_probatoria": p.get("forca_probatoria", "N/D"),
            "categoria": p.get("categoria", "N/D"),
            "proximo_passo": p.get("proximo_passo", "N/D"),
        }
    return result


# ============================================================
# FONTE 2: PRODAM.DB (spcf_faturas + spcf_empenhos)
# ============================================================

def carregar_db() -> dict:
    """Carrega dados agregados de prodam.db."""
    # Copiar para temp dir portável (Windows usa AppData\Local\Temp, Unix /tmp)
    import tempfile
    tmp_db = str(Path(tempfile.gettempdir()) / "reconciliacao_prodam.db")
    shutil.copy(str(DB_PATH), tmp_db)
    conn = sqlite3.connect(tmp_db)
    c = conn.cursor()

    # Faturas por cliente
    c.execute("""
        SELECT cliente,
               COUNT(*) as qtd,
               SUM(valor_bruto) as total_bruto,
               COUNT(CASE WHEN situacao IN ('Emitida','Suspensa','Suspenso') THEN 1 END) as qtd_abertas
        FROM spcf_faturas
        WHERE cliente != ''
        GROUP BY cliente
    """)
    faturas = {}
    for r in c.fetchall():
        faturas[r[0]] = {
            "qtd_faturas": r[1],
            "total_bruto": Decimal(str(r[2])) if r[2] else Decimal("0"),
            "qtd_abertas": r[3] or 0,
        }

    # Empenhos por cliente
    c.execute("""
        SELECT cliente,
               COUNT(*) as qtd,
               SUM(valor) as total_valor
        FROM spcf_empenhos
        WHERE cliente != ''
        GROUP BY cliente
    """)
    empenhos = {}
    for r in c.fetchall():
        empenhos[r[0]] = {
            "qtd_empenhos": r[1],
            "total_empenhos": Decimal(str(r[2])) if r[2] else Decimal("0"),
        }

    # Cruzamento pendrive
    c.execute("""
        SELECT devedor, COUNT(*) as qtd
        FROM cruzamento_spcf_pendrive
        WHERE devedor != ''
        GROUP BY devedor
    """)
    cruzamento = {r[0]: r[1] for r in c.fetchall()}

    conn.close()
    return {"faturas": faturas, "empenhos": empenhos, "cruzamento": cruzamento}


# ============================================================
# FONTE 3: SPCF_EXTRACAO/por_devedor (JSONs)
# ============================================================

def carregar_spcf_json(sigla_dir: str) -> dict:
    """Carrega dados SPCF JSON de um devedor."""
    pasta = SPCF_POR_DEVEDOR / sigla_dir
    if not pasta.is_dir():
        return {}

    result = {}

    # Faturas
    for nome in ["consolidado_faturas", "cobrancas"]:
        arq = pasta / f"{nome}_{sigla_dir}.json"
        if arq.exists():
            with open(arq, "r", encoding="utf-8") as f:
                try:
                    data = json.load(f)
                except json.JSONDecodeError:
                    continue
            if isinstance(data, list):
                total = Decimal("0")
                for item in data:
                    val_str = item.get("Valor") or item.get("dados_base", {}).get("Valor", "0")
                    val = normalizar_brl(str(val_str))
                    if val:
                        total += val
                key = f"{nome}_qtd"
                result[key] = len(data)
                result[f"{nome}_total"] = total

    # Empenhos
    arq_emp = pasta / f"consolidado_empenhos_{sigla_dir}.json"
    if arq_emp.exists():
        with open(arq_emp, "r", encoding="utf-8") as f:
            try:
                data = json.load(f)
            except json.JSONDecodeError:
                data = []
        if isinstance(data, list):
            total = Decimal("0")
            for item in data:
                val_str = item.get("Valor") or item.get("dados_base", {}).get("Valor", "0")
                val = normalizar_brl(str(val_str))
                if val:
                    total += val
            result["empenhos_qtd"] = len(data)
            result["empenhos_total"] = total

    # Contratos
    arq_cont = pasta / f"consolidado_contratos_{sigla_dir}.json"
    if arq_cont.exists():
        with open(arq_cont, "r", encoding="utf-8") as f:
            try:
                data = json.load(f)
            except json.JSONDecodeError:
                data = []
        if isinstance(data, list):
            result["contratos_qtd"] = len(data)

    return result


# ============================================================
# FONTE 4: DOSSIÊ / INVENTARIO.xlsx
# ============================================================

def carregar_dossie(sigla_dossie: str) -> dict:
    """Tenta carregar dados do INVENTARIO.xlsx de um dossiê."""
    pasta = DOSSIE_DIR / f"{sigla_dossie}_DOSSIE"
    inv = pasta / "INVENTARIO.xlsx"

    result = {"dossie_existe": pasta.is_dir(), "inventario_existe": inv.exists()}

    if inv.exists():
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(inv), read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                result[f"sheet_{sheet_name}_linhas"] = len(rows)
                # Tentar contar documentos
                if rows:
                    headers = [str(h).lower() if h else "" for h in rows[0]]
                    result[f"sheet_{sheet_name}_colunas"] = headers
            wb.close()
        except Exception as e:
            result["inventario_erro"] = str(e)

    return result


# ============================================================
# RECONCILIAÇÃO PRINCIPAL
# ============================================================

def reconciliar():
    """Executa a reconciliação completa."""
    print("=" * 60)
    print("RECONCILIAÇÃO 4 FONTES — Projeto PRODAM")
    print(f"Data: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("=" * 60)

    # Carregar fontes
    print("\n[1/4] Carregando profiles.json...")
    profiles = carregar_profiles()
    print(f"       {len(profiles)} devedores")

    print("[2/4] Carregando prodam.db...")
    db = carregar_db()
    print(f"       {len(db['faturas'])} clientes com faturas, {len(db['empenhos'])} com empenhos")

    print("[3/4] Carregando SPCF JSONs...")
    # (carregado por devedor no loop)

    print("[4/4] Verificando dossiês...")
    # (carregado por devedor no loop)

    # Reconciliar por devedor
    resultados = {}
    alertas_criticos = []

    for sigla, prof in sorted(profiles.items(), key=lambda x: x[1]["val_exig"], reverse=True):
        r = {"sigla": sigla, "profiles": prof}

        # === FONTE 2: DB ===
        # Buscar siglas equivalentes no banco
        db_siglas = SIGLA_MAP_DB_FATURAS.get(sigla, [sigla])
        db_fat_total = Decimal("0")
        db_fat_qtd = 0
        db_emp_total = Decimal("0")
        db_emp_qtd = 0

        for ds in db_siglas:
            if ds in db["faturas"]:
                db_fat_total += db["faturas"][ds]["total_bruto"]
                db_fat_qtd += db["faturas"][ds]["qtd_faturas"]
            if ds in db["empenhos"]:
                db_emp_total += db["empenhos"][ds]["total_empenhos"]
                db_emp_qtd += db["empenhos"][ds]["qtd_empenhos"]

        r["db"] = {
            "siglas_usadas": db_siglas,
            "faturas_qtd": db_fat_qtd,
            "faturas_total": db_fat_total,
            "empenhos_qtd": db_emp_qtd,
            "empenhos_total": db_emp_total,
        }

        # === FONTE 3: SPCF JSON ===
        spcf_dir = SIGLA_MAP_SPCF_DIR.get(sigla, sigla)
        if spcf_dir and (SPCF_POR_DEVEDOR / spcf_dir).is_dir():
            r["spcf_json"] = carregar_spcf_json(spcf_dir)
            r["spcf_json"]["pasta"] = spcf_dir
        else:
            r["spcf_json"] = {"pasta": None}

        # === FONTE 4: DOSSIÊ ===
        dossie_siglas = SIGLA_MAP_DOSSIE.get(sigla, [sigla.replace("/", "_").replace(" ", "_")])
        dossie_info = {"encontrado": False}
        for ds in dossie_siglas:
            info = carregar_dossie(ds)
            if info.get("dossie_existe"):
                dossie_info = info
                dossie_info["sigla_dossie"] = ds
                dossie_info["encontrado"] = True
                break
        r["dossie"] = dossie_info

        # === ANÁLISE DE DIVERGÊNCIAS ===
        divergencias = []

        # 1. Faturas: profiles vs db
        prof_fat_total = prof["faturas_total"]
        if db_fat_qtd > 0 and prof_fat_total > 0:
            diff_fat = abs(db_fat_qtd - prof_fat_total)
            if diff_fat > 0:
                divergencias.append({
                    "tipo": "QTD_FATURAS",
                    "fontes": "profiles.json vs prodam.db",
                    "profiles": prof_fat_total,
                    "db": db_fat_qtd,
                    "diff": diff_fat,
                    "nota": "DB pode incluir faturas de outros status"
                })

        # 2. Valores: val_orig (profiles) vs total_bruto (db)
        if db_fat_total > 0 and prof["val_orig"] > 0:
            diff_pct = pct_diff(db_fat_total, prof["val_orig"])
            if diff_pct and diff_pct > 5:
                divergencias.append({
                    "tipo": "VALOR_ORIGINAL",
                    "fontes": "profiles.json(val_orig) vs prodam.db(valor_bruto)",
                    "profiles_val_orig": str(prof["val_orig"]),
                    "db_total_bruto": str(db_fat_total),
                    "diff_pct": round(diff_pct, 1),
                    "nota": "Diferença > 5% — verificar filtros de prescrição"
                })

        # 3. SPCF JSON vs DB
        spcf_fat_qtd = r["spcf_json"].get("consolidado_faturas_qtd") or r["spcf_json"].get("cobrancas_qtd", 0)
        if spcf_fat_qtd and db_fat_qtd and spcf_fat_qtd != db_fat_qtd:
            divergencias.append({
                "tipo": "QTD_FATURAS_SPCF_VS_DB",
                "fontes": "SPCF_JSON vs prodam.db",
                "spcf_json": spcf_fat_qtd,
                "db": db_fat_qtd,
                "diff": abs(spcf_fat_qtd - db_fat_qtd),
            })

        # 4. Dossiê sem inventário
        if not dossie_info.get("encontrado"):
            divergencias.append({
                "tipo": "SEM_DOSSIE",
                "fontes": "PRODAM_DOCS",
                "nota": f"Nenhuma pasta *_DOSSIE encontrada para {sigla}"
            })
        elif not dossie_info.get("inventario_existe"):
            divergencias.append({
                "tipo": "SEM_INVENTARIO",
                "fontes": "PRODAM_DOCS",
                "nota": f"Pasta dossiê existe mas sem INVENTARIO.xlsx"
            })

        r["divergencias"] = divergencias

        # Alertas críticos (divergência > 20% em valor)
        for d in divergencias:
            if d["tipo"] == "VALOR_ORIGINAL" and d.get("diff_pct", 0) > 20:
                alertas_criticos.append({
                    "devedor": sigla,
                    "diff_pct": d["diff_pct"],
                    "profiles": d["profiles_val_orig"],
                    "db": d["db_total_bruto"],
                })

        resultados[sigla] = r

    return resultados, alertas_criticos


# ============================================================
# RELATÓRIO MARKDOWN
# ============================================================

def gerar_relatorio_md(resultados: dict, alertas: list):
    """Gera relatório MD."""
    lines = []
    lines.append("# Reconciliação 4 Fontes — Projeto PRODAM")
    lines.append(f"\n**Data:** {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    lines.append(f"**Devedores analisados:** {len(resultados)}")
    lines.append("")

    # Resumo geral
    total_profiles = sum(r["profiles"]["val_exig"] for r in resultados.values())
    total_db = sum(r["db"]["faturas_total"] for r in resultados.values())
    total_div = sum(len(r["divergencias"]) for r in resultados.values())
    sem_dossie = sum(1 for r in resultados.values() if not r["dossie"].get("encontrado"))

    lines.append("## Resumo Executivo")
    lines.append("")
    lines.append(f"| Métrica | Valor |")
    lines.append(f"|---------|-------|")
    lines.append(f"| Total exigível (profiles.json) | {fmt_brl(total_profiles)} |")
    lines.append(f"| Total bruto (prodam.db) | {fmt_brl(total_db)} |")
    lines.append(f"| Divergências detectadas | {total_div} |")
    lines.append(f"| Devedores sem dossiê | {sem_dossie} |")
    lines.append(f"| Alertas críticos (>20% diff) | {len(alertas)} |")
    lines.append("")

    # Alertas críticos
    if alertas:
        lines.append("## ALERTAS CRÍTICOS (divergência > 20%)")
        lines.append("")
        lines.append("| Devedor | profiles.json (val_orig) | prodam.db (bruto) | Diff % |")
        lines.append("|---------|------------------------|-------------------|--------|")
        for a in sorted(alertas, key=lambda x: x["diff_pct"], reverse=True):
            p = fmt_brl(Decimal(a["profiles"]))
            d = fmt_brl(Decimal(a["db"]))
            lines.append(f"| {a['devedor']} | {p} | {d} | {a['diff_pct']:.1f}% |")
        lines.append("")

    # Tabela consolidada
    lines.append("## Tabela Consolidada por Devedor")
    lines.append("")
    lines.append("| Devedor | Cat. | Val Exig (profiles) | Val Orig (profiles) | Bruto (DB) | Fat Prof | Fat DB | SPCF JSON | Dossiê | Diverg. |")
    lines.append("|---------|------|--------------------|--------------------|------------|----------|--------|-----------|--------|---------|")

    for sigla, r in sorted(resultados.items(), key=lambda x: x[1]["profiles"]["val_exig"], reverse=True):
        p = r["profiles"]
        db = r["db"]
        spcf_fat = r["spcf_json"].get("consolidado_faturas_qtd") or r["spcf_json"].get("cobrancas_qtd", "-")
        dossie_mark = "OK" if r["dossie"].get("encontrado") else "FALTA"
        ndiv = len(r["divergencias"])
        div_mark = f"**{ndiv}**" if ndiv > 0 else "0"

        lines.append(
            f"| {sigla} | {p['categoria'][:3]} | {fmt_brl(p['val_exig'])} | {fmt_brl(p['val_orig'])} "
            f"| {fmt_brl(db['faturas_total'])} | {p['faturas_total']} | {db['faturas_qtd']} "
            f"| {spcf_fat} | {dossie_mark} | {div_mark} |"
        )
    lines.append("")

    # Detalhamento por devedor com divergências
    devedores_com_div = {s: r for s, r in resultados.items() if r["divergencias"]}
    if devedores_com_div:
        lines.append("## Detalhamento de Divergências")
        lines.append("")
        for sigla, r in sorted(devedores_com_div.items(), key=lambda x: x[1]["profiles"]["val_exig"], reverse=True):
            lines.append(f"### {sigla}")
            lines.append("")
            for d in r["divergencias"]:
                tipo = d["tipo"]
                fontes = d.get("fontes", "")
                nota = d.get("nota", "")
                if tipo == "QTD_FATURAS":
                    lines.append(f"- **{tipo}** ({fontes}): profiles={d['profiles']}, db={d['db']} (diff={d['diff']}). {nota}")
                elif tipo == "VALOR_ORIGINAL":
                    lines.append(f"- **{tipo}** ({fontes}): profiles={fmt_brl(Decimal(d['profiles_val_orig']))}, db={fmt_brl(Decimal(d['db_total_bruto']))} (diff={d['diff_pct']:.1f}%). {nota}")
                elif tipo in ("SEM_DOSSIE", "SEM_INVENTARIO"):
                    lines.append(f"- **{tipo}**: {nota}")
                else:
                    lines.append(f"- **{tipo}** ({fontes}): {json.dumps(d, ensure_ascii=False, default=str)}")
            lines.append("")

    # Legenda
    lines.append("## Legenda")
    lines.append("")
    lines.append("- **Val Exig**: Valor exigível (não prescrito) segundo profiles.json")
    lines.append("- **Val Orig**: Valor original histórico (sem correção)")
    lines.append("- **Bruto (DB)**: Soma de valor_bruto de todas as faturas no prodam.db (inclui prescritas)")
    lines.append("- **Fat Prof / Fat DB**: Contagem de faturas no profiles.json vs prodam.db")
    lines.append("- **SPCF JSON**: Faturas encontradas nos JSONs de SPCF_EXTRACAO/por_devedor/")
    lines.append("- **Dossiê**: Se existe pasta *_DOSSIE com INVENTARIO.xlsx")
    lines.append("")
    lines.append("### Fontes de Dados")
    lines.append("")
    lines.append("1. **profiles.json** (SSOT) — Fonte autoritativa com valores consolidados por devedor")
    lines.append("2. **prodam.db** — Banco SQLite com faturas e empenhos importados do SPCF")
    lines.append("3. **SPCF_EXTRACAO/por_devedor/** — JSONs crus extraídos do sistema SPCF por web scraping")
    lines.append("4. **PRODAM_DOCS/*_DOSSIE/** — Dossiês documentais com INVENTARIO.xlsx dos PDFs originais")

    return "\n".join(lines)


# ============================================================
# JSON EXPORT
# ============================================================

def serializar_resultado(resultados: dict) -> dict:
    """Serializa para JSON (Decimal → str)."""
    def _convert(obj):
        if isinstance(obj, Decimal):
            return str(obj)
        if isinstance(obj, set):
            return list(obj)
        return obj

    return json.loads(json.dumps(resultados, default=_convert, ensure_ascii=False))


# ============================================================
# MAIN
# ============================================================

def main():
    resultados, alertas = reconciliar()

    # Relatório MD
    md = gerar_relatorio_md(resultados, alertas)
    md_path = OUTPUT_DIR / "reconciliacao_4_fontes_RELATORIO.md"
    with open(md_path, "w", encoding="utf-8") as f:
        f.write(md)
    print(f"\n✓ Relatório MD: {md_path}")

    # JSON
    json_data = serializar_resultado(resultados)
    json_path = OUTPUT_DIR / "reconciliacao_4_fontes_DADOS.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(json_data, f, ensure_ascii=False, indent=2)
    print(f"✓ Dados JSON: {json_path}")

    # Resumo rápido
    total_div = sum(len(r["divergencias"]) for r in resultados.values())
    print(f"\n{'='*60}")
    print(f"RESULTADO: {len(resultados)} devedores analisados")
    print(f"  Divergências: {total_div}")
    print(f"  Alertas críticos: {len(alertas)}")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
