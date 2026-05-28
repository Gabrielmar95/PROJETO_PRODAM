"""
dossie_multiformato_devedor.py — Gera dossiê completo em 5 formatos por devedor.

Para cada devedor do profiles.json, produz:
  DOSSIES_MULTIFORMATO/<SIGLA>/
    ├── dossie.md         — relatório humano markdown
    ├── dossie.html       — relatório interativo Chart.js
    ├── dossie.xlsx       — planilha multi-aba (requer openpyxl)
    ├── dossie.json       — dados brutos estruturados
    └── csv/
        ├── resumo.csv
        ├── empenhos.csv
        ├── faturas.csv
        └── cobrancas.csv

Uso:
  py -3.12 dossie_multiformato_devedor.py                # todos devedores
  py -3.12 dossie_multiformato_devedor.py --devedor SES  # um devedor
"""

from __future__ import annotations
import json, sys, sqlite3, csv, re, argparse
from pathlib import Path
from datetime import date

sys.stdout.reconfigure(encoding="utf-8")

# Helpers compartilhados — norm() resolve POLÍCIA CIVIL vs POLICIA CIVIL
ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(ROOT / "scripts"))
from prodam_utils import norm, norm_variants, brl, fmt_brl
import os
if os.environ.get("PRODAM_FREEZE_EMISSAO"):
    sys.exit("[FREEZE] Emissão de peças bloqueada durante auditoria DE. Remover PRODAM_FREEZE_EMISSAO para destravar.")

DB = sqlite3.connect(str(ROOT / "prodam.db"))
DOCS = ROOT / "PRODAM_DOCS"
DADOS = ROOT / "SPCF_EXTRACAO" / "dados"
OUT = ROOT / "DOSSIES_MULTIFORMATO"
OUT.mkdir(exist_ok=True)

HOJE = date.today()


def sigla_to_folder(sigla: str) -> str:
    return re.sub(r"[^\w]", "_", sigla).strip("_")


# ============================================================
# CARREGA FONTES
# ============================================================
print("Carregando fontes...")
profiles = json.load(open(DOCS / "profiles.json", encoding="utf-8"))
con_f = json.load(open(DADOS / "consolidado_faturas.json", encoding="utf-8"))
cobrancas = json.load(open(DADOS / "cobrancas.json", encoding="utf-8"))
print(
    f"  {len(profiles)} devedores, {len(con_f)} faturas SPCF, {len(cobrancas)} cobranças"
)


# ============================================================
# DADOS POR DEVEDOR
# ============================================================
def _match_clientes(sigla: str) -> list:
    """Retorna lista de clientes do DB que casam com a sigla via norm()."""
    variantes = norm_variants(sigla)
    all_clients = [
        r[0]
        for r in DB.execute(
            "SELECT DISTINCT cliente FROM spcf_empenhos UNION SELECT DISTINCT cliente FROM spcf_faturas"
        ).fetchall()
        if r[0]
    ]
    return [
        c
        for c in all_clients
        if norm(c) in variantes or any(v in norm(c) for v in variantes)
    ]


def coletar_dados(sigla: str, profile: dict) -> dict:
    # Use norm() para pegar todas as variantes do cliente (POLÍCIA CIVIL vs POLICIA CIVIL)
    clientes_matched = _match_clientes(sigla)
    if not clientes_matched:
        # Fallback: busca LIKE legada
        sigla_q = sigla.upper()
        sigla_base = sigla_q.split("/")[0]
        placeholders = (f"%{sigla_q}%", f"%{sigla_base}%")
        empenhos = DB.execute(
            """
            SELECT numero, contrato_ref, valor, situacao, data_emissao,
                   json_extract(dados_base, '$.Descrição') as descricao
            FROM spcf_empenhos
            WHERE UPPER(cliente) LIKE ? OR UPPER(cliente) LIKE ?
            ORDER BY data_emissao DESC
        """,
            placeholders,
        ).fetchall()
        faturas = DB.execute(
            """
            SELECT id, nf, contrato_num, valor_bruto, competencia, situacao,
                   cadeia_completude, total_empenhos_vinculados
            FROM spcf_faturas
            WHERE UPPER(cliente) LIKE ? OR UPPER(cliente) LIKE ?
            ORDER BY valor_bruto DESC
        """,
            placeholders,
        ).fetchall()
    else:
        # Busca exata pelos clientes casados via norm()
        cond = " OR ".join("cliente = ?" for _ in clientes_matched)
        empenhos = DB.execute(
            f"""
            SELECT numero, contrato_ref, valor, situacao, data_emissao,
                   json_extract(dados_base, '$.Descrição') as descricao
            FROM spcf_empenhos WHERE {cond}
            ORDER BY data_emissao DESC
        """,
            clientes_matched,
        ).fetchall()
        faturas = DB.execute(
            f"""
            SELECT id, nf, contrato_num, valor_bruto, competencia, situacao,
                   cadeia_completude, total_empenhos_vinculados
            FROM spcf_faturas WHERE {cond}
            ORDER BY valor_bruto DESC
        """,
            clientes_matched,
        ).fetchall()

    # 3. Cobranças (em cobrancas.json) — usa norm()
    variantes_cob = norm_variants(sigla)
    cobr = [
        r
        for r in cobrancas
        if norm(r.get("Cliente", "")) in variantes_cob
        or any(v in norm(r.get("Cliente", "")) for v in variantes_cob)
    ]

    # 4. Agregações
    emp_stats = {
        "qtd": len(empenhos),
        "valor_total": sum(float(e[2] or 0) for e in empenhos),
        "por_ano": {},
    }
    for e in empenhos:
        data = e[4] or ""
        ano = data[-4:] if len(data) >= 10 else "?"
        emp_stats["por_ano"].setdefault(ano, {"qtd": 0, "valor": 0})
        emp_stats["por_ano"][ano]["qtd"] += 1
        emp_stats["por_ano"][ano]["valor"] += float(e[2] or 0)

    fat_stats = {
        "qtd": len(faturas),
        "valor_total": sum(float(f[3] or 0) for f in faturas),
    }

    cobr_stats = {
        "qtd": len(cobr),
        "valor_total": float(sum(brl(c.get("Valor", "0")) for c in cobr)),
    }

    return {
        "sigla": sigla,
        "profile": profile,
        "empenhos": [
            {
                "numero": e[0],
                "contrato": e[1],
                "valor": float(e[2] or 0),
                "situacao": e[3],
                "data_emissao": e[4],
                "descricao": e[5],
            }
            for e in empenhos
        ],
        "faturas": [
            {
                "id": f[0],
                "nf": f[1],
                "contrato": f[2],
                "valor_bruto": float(f[3] or 0),
                "competencia": f[4],
                "situacao": f[5],
                "cadeia": f[6],
                "empenhos_vinc": f[7],
            }
            for f in faturas
        ],
        "cobrancas": cobr,
        "stats": {"empenhos": emp_stats, "faturas": fat_stats, "cobrancas": cobr_stats},
    }


# ============================================================
# GERADORES DE FORMATO
# ============================================================
def gerar_md(dados: dict, pasta: Path):
    p = dados["profile"]
    md = f"""# Dossiê Completo — {dados['sigla']}
**{p.get('nome_completo','')}** | Categoria: **{p.get('categoria','')}**
Data: {HOJE.isoformat()} | CNPJ: {p.get('cnpj','—')}

## Visão Executiva

| Métrica | Valor |
|---------|-------|
| Força probatória | {p.get('forca_probatoria','—')} |
| Próximo passo | {p.get('proximo_passo','—')} |
| Fase atual | {p.get('fase_atual','—')} |
| Regime de execução | {p.get('regime_execucao','—')} |
| Índice correção | {p.get('indice_correcao','—')} |
| Juros mora | {p.get('juros_mora','—')} |
| Valor exigível (profile) | {fmt_brl(p.get('val_exig') or 0)} |
| Valor atualizado (profile) | {fmt_brl(p.get('val_atualizado') or 0)} |
| Faturas total | {p.get('faturas_total',0)} |
| Faturas exigíveis | {p.get('faturas_exigiveis',0)} |
| Faturas prescritas | {p.get('faturas_prescritas',0)} |
| Probabilidade recuperação | {p.get('p_recuperacao',0):.0%} |
| Valor esperado (E[V]) | {fmt_brl(p.get('ev_valor_esperado') or 0)} |
| Honorários esperados | {fmt_brl(p.get('ev_honorarios') or 0)} |

## Dados SPCF (prodam.db)

### Empenhos: {dados['stats']['empenhos']['qtd']} NEs | {fmt_brl(dados['stats']['empenhos']['valor_total'])}

Distribuição por ano:

| Ano | Qtd | Valor |
|-----|-----|-------|
"""
    for ano in sorted(dados["stats"]["empenhos"]["por_ano"].keys()):
        s = dados["stats"]["empenhos"]["por_ano"][ano]
        md += f"| {ano} | {s['qtd']} | {fmt_brl(s['valor'])} |\n"

    md += f"""
### Faturas: {dados['stats']['faturas']['qtd']} | {fmt_brl(dados['stats']['faturas']['valor_total'])}
### Cobranças SPCF: {dados['stats']['cobrancas']['qtd']} | {fmt_brl(dados['stats']['cobrancas']['valor_total'])}

## Top 10 Empenhos (por valor)

| Número | Contrato | Data | Valor | Situação |
|--------|----------|------|-------|----------|
"""
    for e in sorted(dados["empenhos"], key=lambda x: -x["valor"])[:10]:
        md += f"| {e['numero']} | {e['contrato']} | {e['data_emissao']} | {fmt_brl(e['valor'])} | {e['situacao']} |\n"

    md += "\n## Top 10 Faturas (por valor)\n\n| ID | NF | Contrato | Competência | Valor | Situação | Cadeia |\n|-----|----|----|----|----|----|----|\n"
    for f in dados["faturas"][:10]:
        md += f"| {f['id']} | {f['nf']} | {f['contrato']} | {f['competencia']} | {fmt_brl(f['valor_bruto'])} | {f['situacao']} | {f['cadeia']} |\n"

    md += f"""

## Observações do Profile

{p.get('observacoes','—')}

## Alertas de Prescrição

- Data próxima prescrição: **{p.get('data_prescricao_proxima','—')}**
- Reconhecimento revisado: {'✅' if p.get('reconhecimento_revisado') else '❌'}
- Título executivo: {'✅' if p.get('titulo_executivo') else '❌'}

---
_Gerado por `dossie_multiformato_devedor.py` em {HOJE.isoformat()}_
"""
    (pasta / "dossie.md").write_text(md, encoding="utf-8")


def gerar_html(dados: dict, pasta: Path):
    p = dados["profile"]
    por_ano = dados["stats"]["empenhos"]["por_ano"]
    anos = sorted(por_ano.keys())
    valores = [por_ano[a]["valor"] for a in anos]
    qtds = [por_ano[a]["qtd"] for a in anos]

    top_emp = sorted(dados["empenhos"], key=lambda x: -x["valor"])[:10]
    top_fat = dados["faturas"][:10]

    html = f"""<!DOCTYPE html>
<html lang="pt-BR"><head><meta charset="utf-8">
<title>Dossiê {dados['sigla']} — Projeto PRODAM</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
<style>
body{{font-family:Segoe UI,Arial,sans-serif;max-width:1400px;margin:20px auto;padding:20px;background:#f5f5f5}}
h1,h2{{color:#1F3864}} h1{{border-bottom:3px solid #B8963E}}
.grid{{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin:16px 0}}
.card{{background:white;padding:14px;border-radius:6px;box-shadow:0 2px 4px rgba(0,0,0,0.08)}}
.kpi{{font-size:22px;color:#1F3864;font-weight:700}}
.label{{color:#777;font-size:11px;text-transform:uppercase;letter-spacing:0.5px}}
table{{width:100%;border-collapse:collapse;background:white;margin:10px 0}}
th,td{{padding:7px 10px;border-bottom:1px solid #eee;text-align:left;font-size:13px}}
th{{background:#1F3864;color:white}}
tr:hover{{background:#fafafa}}
.forte{{color:green;font-weight:bold}} .media{{color:orange;font-weight:bold}} .fraca{{color:red}}
canvas{{max-height:300px}}
</style></head><body>
<h1>Dossiê Completo — {dados['sigla']}</h1>
<p><b>{p.get('nome_completo','')}</b> | {p.get('categoria','')} | CNPJ: {p.get('cnpj','—')} | Data: {HOJE.isoformat()}</p>

<div class="grid">
<div class="card"><div class="label">Força</div><div class="kpi {p.get('forca_probatoria','').lower()}">{p.get('forca_probatoria','—')}</div></div>
<div class="card"><div class="label">Próximo passo</div><div class="kpi" style="font-size:15px">{p.get('proximo_passo','—')}</div></div>
<div class="card"><div class="label">Val. exigível</div><div class="kpi">{fmt_brl(p.get('val_exig') or 0)}</div></div>
<div class="card"><div class="label">Val. atualizado</div><div class="kpi">{fmt_brl(p.get('val_atualizado') or 0)}</div></div>
<div class="card"><div class="label">Faturas exig.</div><div class="kpi">{p.get('faturas_exigiveis',0)}</div></div>
<div class="card"><div class="label">Faturas prescr.</div><div class="kpi">{p.get('faturas_prescritas',0)}</div></div>
<div class="card"><div class="label">Empenhos (DB)</div><div class="kpi">{dados['stats']['empenhos']['qtd']}</div></div>
<div class="card"><div class="label">Prob. recup.</div><div class="kpi">{p.get('p_recuperacao',0):.0%}</div></div>
</div>

<h2>Empenhos por Ano</h2>
<div class="card"><canvas id="empAno"></canvas></div>

<h2>Top 10 Empenhos (por valor)</h2>
<table><tr><th>Número</th><th>Contrato</th><th>Data</th><th>Valor</th><th>Situação</th></tr>
"""
    for e in top_emp:
        html += f"<tr><td>{e['numero']}</td><td>{e['contrato']}</td><td>{e['data_emissao']}</td><td>{fmt_brl(e['valor'])}</td><td>{e['situacao']}</td></tr>\n"
    html += "</table>\n<h2>Top 10 Faturas</h2>\n<table><tr><th>ID</th><th>NF</th><th>Contrato</th><th>Comp.</th><th>Valor</th><th>Situação</th><th>Cadeia</th></tr>"
    for f in top_fat:
        html += f"<tr><td>{f['id']}</td><td>{f['nf']}</td><td>{f['contrato']}</td><td>{f['competencia']}</td><td>{fmt_brl(f['valor_bruto'])}</td><td>{f['situacao']}</td><td>{f['cadeia']}</td></tr>\n"
    html += f"""</table>

<h2>Observações</h2>
<div class="card"><p>{p.get('observacoes','—')}</p></div>

<script>
new Chart(document.getElementById('empAno'),{{
  type:'bar',
  data:{{labels:{json.dumps(anos)},datasets:[
    {{label:'Qtd',data:{json.dumps(qtds)},yAxisID:'y',backgroundColor:'#1F3864'}},
    {{label:'Valor R$',data:{json.dumps(valores)},yAxisID:'y1',type:'line',borderColor:'#B8963E'}}
  ]}},
  options:{{scales:{{y:{{position:'left'}},y1:{{position:'right',grid:{{drawOnChartArea:false}}}}}}}}
}});
</script>
</body></html>"""
    (pasta / "dossie.html").write_text(html, encoding="utf-8")


def gerar_json(dados: dict, pasta: Path):
    with open(pasta / "dossie.json", "w", encoding="utf-8") as fh:
        json.dump(dados, fh, ensure_ascii=False, indent=2, default=str)


def gerar_csv(dados: dict, pasta: Path):
    csv_dir = pasta / "csv"
    csv_dir.mkdir(exist_ok=True)

    # Resumo
    with open(csv_dir / "resumo.csv", "w", encoding="utf-8-sig", newline="") as fh:
        w = csv.writer(fh, delimiter=";")
        w.writerow(["campo", "valor"])
        p = dados["profile"]
        for k in [
            "nome_completo",
            "categoria",
            "cnpj",
            "forca_probatoria",
            "proximo_passo",
            "val_exig",
            "val_atualizado",
            "faturas_exigiveis",
            "faturas_prescritas",
            "faturas_total",
            "p_recuperacao",
            "indice_correcao",
        ]:
            w.writerow([k, p.get(k, "")])

    # Empenhos
    with open(csv_dir / "empenhos.csv", "w", encoding="utf-8-sig", newline="") as fh:
        w = csv.writer(fh, delimiter=";")
        w.writerow(
            ["numero", "contrato", "data_emissao", "valor", "situacao", "descricao"]
        )
        for e in dados["empenhos"]:
            w.writerow(
                [
                    e["numero"],
                    e["contrato"],
                    e["data_emissao"],
                    str(e["valor"]).replace(".", ","),
                    e["situacao"],
                    (e["descricao"] or "")[:200],
                ]
            )

    # Faturas
    with open(csv_dir / "faturas.csv", "w", encoding="utf-8-sig", newline="") as fh:
        w = csv.writer(fh, delimiter=";")
        w.writerow(
            [
                "id",
                "nf",
                "contrato",
                "competencia",
                "valor_bruto",
                "situacao",
                "cadeia",
                "empenhos_vinc",
            ]
        )
        for f in dados["faturas"]:
            w.writerow(
                [
                    f["id"],
                    f["nf"],
                    f["contrato"],
                    f["competencia"],
                    str(f["valor_bruto"]).replace(".", ","),
                    f["situacao"],
                    f["cadeia"],
                    f["empenhos_vinc"],
                ]
            )

    # Cobranças
    with open(csv_dir / "cobrancas.csv", "w", encoding="utf-8-sig", newline="") as fh:
        w = csv.writer(fh, delimiter=";")
        if dados["cobrancas"]:
            keys = [k for k in dados["cobrancas"][0] if not k.endswith("_link")]
            w.writerow(keys)
            for c in dados["cobrancas"]:
                w.writerow([c.get(k, "") for k in keys])


def gerar_xlsx(dados: dict, pasta: Path):
    try:
        from openpyxl import Workbook
    except ImportError:
        return False
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("Cenarios")
    ws.title = "Resumo"
    p = dados["profile"]
    ws.append(["Campo", "Valor"])
    for k in [
        "sigla",
        "nome_completo",
        "categoria",
        "cnpj",
        "forca_probatoria",
        "proximo_passo",
        "val_exig",
        "val_atualizado",
        "faturas_exigiveis",
        "faturas_prescritas",
        "p_recuperacao",
        "indice_correcao",
        "data_prescricao_proxima",
    ]:
        ws.append([k, str(p.get(k, "") if k != "sigla" else dados["sigla"])])

    ws2 = wb.create_sheet("Empenhos")
    ws2.append(["Numero", "Contrato", "Data", "Valor", "Situacao", "Descricao"])
    for e in dados["empenhos"]:
        ws2.append(
            [
                e["numero"],
                e["contrato"],
                e["data_emissao"],
                float(e["valor"]),
                e["situacao"],
                (e["descricao"] or "")[:200],
            ]
        )

    ws3 = wb.create_sheet("Faturas")
    ws3.append(
        [
            "ID",
            "NF",
            "Contrato",
            "Competencia",
            "ValorBruto",
            "Situacao",
            "Cadeia",
            "EmpenhosVinc",
        ]
    )
    for f in dados["faturas"]:
        ws3.append(
            [
                f["id"],
                f["nf"],
                f["contrato"],
                f["competencia"],
                float(f["valor_bruto"]),
                f["situacao"],
                f["cadeia"],
                f["empenhos_vinc"],
            ]
        )

    ws4 = wb.create_sheet("Cobrancas")
    if dados["cobrancas"]:
        keys = [k for k in dados["cobrancas"][0] if not k.endswith("_link")]
        ws4.append(keys)
        for c in dados["cobrancas"]:
            ws4.append([str(c.get(k, ""))[:500] for k in keys])

    wb.save(pasta / "dossie.xlsx")
    return True


# ============================================================
# MAIN
# ============================================================
def gerar_devedor(sigla: str, profile: dict):
    pasta = OUT / sigla_to_folder(sigla)
    pasta.mkdir(exist_ok=True, parents=True)
    dados = coletar_dados(sigla, profile)
    gerar_md(dados, pasta)
    gerar_html(dados, pasta)
    gerar_json(dados, pasta)
    gerar_csv(dados, pasta)
    xlsx_ok = gerar_xlsx(dados, pasta)
    return {
        "pasta": str(pasta),
        "xlsx": xlsx_ok,
        "empenhos": dados["stats"]["empenhos"]["qtd"],
        "faturas": dados["stats"]["faturas"]["qtd"],
    }


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--devedor", help="Sigla específica (ex: SES/SUSAM)")
    args = ap.parse_args()

    # _metadata não é devedor — pula entradas com prefixo _
    devedores_validos = {k: v for k, v in profiles.items() if not k.startswith("_")}

    if args.devedor:
        if args.devedor not in devedores_validos:
            print(f"Devedor '{args.devedor}' não encontrado. Disponíveis:")
            for s in sorted(devedores_validos.keys()):
                print(f"  - {s}")
            sys.exit(1)
        r = gerar_devedor(args.devedor, devedores_validos[args.devedor])
        print(f"✅ {args.devedor}: {r}")
    else:
        total = len(devedores_validos)
        print(
            f"Gerando dossiês para {total} devedores ({len(profiles)-total} entradas _metadata ignoradas)..."
        )
        ok = 0
        for sigla, p in devedores_validos.items():
            try:
                r = gerar_devedor(sigla, p)
                ok += 1
                print(
                    f"  [{ok}/{total}] {sigla}: {r['empenhos']} NEs, {r['faturas']} fat"
                )
            except Exception as e:
                print(f"  ERRO {sigla}: {e}")
        print(f"\n✅ {ok}/{total} dossiês gerados em {OUT}")
