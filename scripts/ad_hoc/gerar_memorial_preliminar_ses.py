#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""gerar_memorial_preliminar_ses.py — Memorial Preliminar SES/SUSAM (TRD 2026-05-12)

Script ad-hoc para gerar memorial fatura-a-fatura com correção SELIC pura,
fundamentado em norma legal (EC 113/2021 Art. 3º + Tema 1.368 STJ + Art. 406 CC
pós-Lei 14.905/2024), aplicável a administração direta estadual sem
necessidade de extração contratual cláusula a cláusula.

Universo: 82 candidatas a TRD (3 Tier 1 do 74/2021 + 79 Tier 2 ressuscitadas).
Fatura id 122703 (já prescrita D-71) EXCLUÍDA.

Data-base: 2026-05-12. Output em DOCUMENTOS_GERADOS/SES_SUSAM/.
"""
from __future__ import annotations
import datetime as dt
import json
import sqlite3
import sys
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Optional

# ─── Constantes ─────────────────────────────────────────────────────────────
DB_PATH = Path(r"C:\Users\gabri\Desktop\PROJETO_PRODAM\PRODAM_DOCS\_ANALISE\prodam.db")
OUT_DIR = Path(r"C:\Users\gabri\Desktop\PROJETO_PRODAM\DOCUMENTOS_GERADOS\SES_SUSAM")
SKILL_CACHE = Path(r"C:\Users\gabri\.claude\skills\memorial-calculo-prodam\cache")
DATA_BASE = dt.date(2026, 5, 12)
CUTOFF_PRESCRICAO = DATA_BASE - dt.timedelta(days=5*365 + 1)  # 2021-05-12 aprox
EXCLUIR_PRESCRITA = {122703}  # já prescreveu D-71 — fora da TRD
ZERO = Decimal("0")
D2 = Decimal("0.01")
D6 = Decimal("0.000001")
CEM = Decimal("100")
ROUND = ROUND_HALF_UP

# ─── Helpers ────────────────────────────────────────────────────────────────
def to_dec(x) -> Decimal:
    if x is None or x == "":
        return ZERO
    return Decimal(str(x))

def q2(x: Decimal) -> Decimal:
    return x.quantize(D2, rounding=ROUND)

def q6(x: Decimal) -> Decimal:
    return x.quantize(D6, rounding=ROUND)

def fmt_brl(x: Decimal) -> str:
    s = f"{q2(x):,.2f}"
    return s.replace(",", "X").replace(".", ",").replace("X", ".")

def parse_competencia(s: str) -> dt.date:
    mes, ano = s.split("/")
    return dt.date(int(ano), int(mes), 1)

def parse_dmy(s: str) -> Optional[dt.date]:
    if not s:
        return None
    try:
        return dt.datetime.strptime(s, "%d/%m/%Y").date()
    except (ValueError, TypeError):
        return None

# ─── SELIC fator BCB série 4390 ─────────────────────────────────────────────
def fator_selic_acumulado(ini: dt.date, fim: dt.date) -> Decimal:
    """Acumula (1 + selic_mensal/100) entre ini e fim. Reusa cache da skill."""
    if ini >= fim:
        return Decimal("1")
    cache_p = SKILL_CACHE / f"bcb_4390_{ini.isoformat()}_{fim.isoformat()}.json"
    if cache_p.exists():
        age_hours = (dt.datetime.now() - dt.datetime.fromtimestamp(cache_p.stat().st_mtime)).total_seconds() / 3600
        if age_hours < 24:
            try:
                return Decimal(json.loads(cache_p.read_text(encoding="utf-8"))["fator"])
            except Exception:
                pass
    import requests
    url = "https://api.bcb.gov.br/dados/serie/bcdata.sgs.4390/dados"
    params = {
        "formato": "json",
        "dataInicial": ini.strftime("%d/%m/%Y"),
        "dataFinal": fim.strftime("%d/%m/%Y"),
    }
    resp = requests.get(url, params=params, timeout=30)
    resp.raise_for_status()
    fator = Decimal("1")
    for item in resp.json():
        v = to_dec(item["valor"])
        fator *= (Decimal("1") + v / CEM)
    fator = q6(fator)
    try:
        SKILL_CACHE.mkdir(parents=True, exist_ok=True)
        cache_p.write_text(json.dumps({
            "serie": 4390,
            "data_inicial": ini.isoformat(),
            "data_final": fim.isoformat(),
            "fator": str(fator),
            "fonte": url,
            "fetched_at": dt.datetime.now().isoformat(),
        }, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception:
        pass
    return fator

# ─── Seleção das 82 faturas candidatas ─────────────────────────────────────
def selecionar_candidatas() -> list[dict]:
    """Aplica a mesma lógica da Etapa 1.7 do plano TRD:

       - Filtra spcf_faturas WHERE cliente='SES' AND status válido
       - Para cada fatura, calcula vencimento = competencia + 30 dias
       - Tier 1: vencimento + 5a > hoje (Art. 206 §5º I CC, prazo natural)
       - Tier 2: existe NE do mesmo contrato com data_emissao > vencimento_fatura
                 AND data_emissao >= cutoff_prescricao (Art. 202 VI CC)
       - Exclui fatura id 122703 (já prescrita D-71)
    """
    con = sqlite3.connect(str(DB_PATH))
    con.row_factory = sqlite3.Row
    faturas = con.execute("""
        SELECT id, nf, contrato_num, competencia, valor_bruto, situacao
        FROM spcf_faturas
        WHERE cliente = 'SES'
          AND competencia IS NOT NULL
          AND contrato_num IS NOT NULL
    """).fetchall()

    empenhos = con.execute("""
        SELECT contrato_ref AS contrato, numero, data_emissao
        FROM spcf_empenhos
        WHERE cliente = 'SES'
          AND data_emissao IS NOT NULL
    """).fetchall()
    con.close()

    # Indexar empenhos por contrato → lista de (data_emissao_date, numero)
    # spcf_empenhos.contrato_ref vem em formato variável ('14/2019', '014/2019'),
    # então normalizamos removendo zero-pad para casar com spcf_faturas.contrato_num.
    def norm_ctr(s) -> str:
        if not s:
            return ""
        s = str(s).strip()
        # Remove prefixos de modalidade comuns ("P.", "C.", "T.", "TC.", "CT.")
        for pref in ("CT.", "TC.", "P.", "C.", "T."):
            if s.upper().startswith(pref):
                s = s[len(pref):].strip()
                break
        if "/" in s:
            num, ano = s.split("/", 1)
            return f"{int(num)}/{ano}" if num.isdigit() else s
        return s
    ne_por_contrato: dict[str, list[tuple[dt.date, str]]] = {}
    for ne in empenhos:
        d = parse_dmy(ne["data_emissao"])
        if d:
            ne_por_contrato.setdefault(norm_ctr(ne["contrato"]), []).append((d, ne["numero"]))
    for c in ne_por_contrato:
        ne_por_contrato[c].sort(key=lambda t: t[0])

    candidatas = []
    for f in faturas:
        fid = f["id"]
        if fid in EXCLUIR_PRESCRITA:
            continue
        try:
            comp = parse_competencia(f["competencia"])
        except Exception:
            continue
        # vencimento contábil = primeiro dia do mês seguinte (convenção SES: comp + 30d)
        venc = comp + dt.timedelta(days=30)
        prazo_natural_5a = venc + dt.timedelta(days=5*365 + 1)
        ctr = norm_ctr(f["contrato_num"])
        valor = to_dec(f["valor_bruto"])
        situacao = (f["situacao"] or "").strip()

        # Tier 1: prazo natural ainda não consumado
        tier1 = prazo_natural_5a > DATA_BASE

        # Tier 2: NE pós-vencimento dentro do cutoff
        nes_pos = [ne for ne in ne_por_contrato.get(ctr, []) if ne[0] > venc and ne[0] >= CUTOFF_PRESCRICAO]
        tier2 = bool(nes_pos)

        if not (tier1 or tier2):
            continue
        # Excluir paga / cancelada
        if "paga" in situacao.lower() or "cancel" in situacao.lower():
            # exceto "parcialmente paga"
            if "parcialmente" not in situacao.lower():
                continue

        ultima_ne = nes_pos[-1] if nes_pos else None
        if tier1 and not tier2:
            tier = "1"
            prazo_critico = prazo_natural_5a
        elif tier2:
            tier = "2"
            # prazo = última NE + 2,5a (Decreto 20.910/1932)
            prazo_critico = ultima_ne[0] + dt.timedelta(days=int(2.5*365.25))
        else:
            continue

        candidatas.append({
            "id": fid,
            "nf": f["nf"] or "",
            "contrato": ctr,
            "competencia": comp,
            "vencimento": venc,
            "valor_bruto": valor,
            "situacao": situacao,
            "tier": tier,
            "ultima_ne_pos_venc": ultima_ne[1] if ultima_ne else None,
            "ultima_ne_data": ultima_ne[0] if ultima_ne else None,
            "prazo_critico": prazo_critico,
        })

    return candidatas

# ─── Cálculo SELIC fatura-a-fatura ─────────────────────────────────────────
def calcular_memorial(candidatas: list[dict]) -> tuple[list[dict], dict]:
    print(f"Aplicando SELIC série BCB 4390 a {len(candidatas)} faturas (data-base {DATA_BASE.isoformat()})...")
    resultados = []
    for f in candidatas:
        fator = fator_selic_acumulado(f["vencimento"], DATA_BASE)
        valor_atualizado = q2(f["valor_bruto"] * fator)
        correcao_juros_selic = q2(valor_atualizado - f["valor_bruto"])
        resultados.append({
            **f,
            "fator_selic": fator,
            "valor_atualizado": valor_atualizado,
            "correcao_juros_selic": correcao_juros_selic,
            "meses_entre": (DATA_BASE.year - f["vencimento"].year) * 12 + (DATA_BASE.month - f["vencimento"].month),
        })

    total_bruto = sum((r["valor_bruto"] for r in resultados), ZERO)
    total_atualizado = sum((r["valor_atualizado"] for r in resultados), ZERO)
    total_correcao = q2(total_atualizado - total_bruto)
    tier1 = [r for r in resultados if r["tier"] == "1"]
    tier2 = [r for r in resultados if r["tier"] == "2"]
    totais = {
        "n_total": len(resultados),
        "n_tier1": len(tier1),
        "n_tier2": len(tier2),
        "principal_bruto": total_bruto,
        "tier1_principal_bruto": sum((r["valor_bruto"] for r in tier1), ZERO),
        "tier2_principal_bruto": sum((r["valor_bruto"] for r in tier2), ZERO),
        "correcao_juros_selic": total_correcao,
        "total_atualizado": total_atualizado,
        "honorarios_20pct": q2(total_atualizado * Decimal("0.20")),
        "fator_medio": q6(total_atualizado / total_bruto) if total_bruto else Decimal("0"),
    }
    return resultados, totais

# ─── Render Markdown ───────────────────────────────────────────────────────
def render_md(resultados: list[dict], totais: dict) -> str:
    L = []
    L.append("# Memorial Preliminar de Cálculo — SES/SUSAM")
    L.append("")
    L.append("**Contrato 002/2026 — PRODAM S.A. × Brandão Ozores Advogados**")
    L.append(f"**Data-base:** {DATA_BASE.strftime('%d/%m/%Y')}  ")
    L.append(f"**Devedor:** Secretaria de Saúde do Amazonas / Fundação de Medicina Tropical (SES/SUSAM)  ")
    L.append(f"**CNPJ:** 00.697.295/0001-05  ")
    L.append(f"**Faturas candidatas a TRD:** {totais['n_total']} (Tier 1: {totais['n_tier1']} | Tier 2: {totais['n_tier2']})  ")
    L.append("")
    L.append("---")
    L.append("")
    L.append("## 1. Fundamentação da Atualização Monetária")
    L.append("")
    L.append("A SES/SUSAM é órgão da **administração direta estadual** do Amazonas. Para débitos da Fazenda Pública estadual, a atualização monetária e os juros de mora seguem regime estabelecido em **norma legal** — não em cláusula contratual — conforme:")
    L.append("")
    L.append("- **EC 113/2021, Art. 3º** — Para todos os créditos contra a Fazenda Pública, incidência única da **taxa SELIC** acumulada mensalmente, englobando correção monetária e juros de mora.")
    L.append("- **Art. 406 do Código Civil** (redação dada pela Lei nº 14.905/2024) — Confirma a SELIC como taxa legal aplicável quando inexistente cláusula contratual específica.")
    L.append("- **REsp 793.969/RJ** (1ª Turma do STJ, Rel. p/ acórdão Min. José Delgado, j. 21/02/2006) — Composição documental (Contrato administrativo + Nota de Empenho + Nota Fiscal + Atesto) constitui título executivo dotado de liquidez, certeza e exigibilidade, aplicável à atualização monetária de débitos públicos reconhecidos.")
    L.append("")
    L.append("Como a SELIC já embute correção monetária + juros (Art. 406 CC), não há juros separados nem multa moratória.")
    L.append("")
    L.append("**Série utilizada:** Banco Central do Brasil — código **4390** (SELIC mensal acumulada, valor efetivo em % a.m.).")
    L.append("**Política de arredondamento:** ROUND_HALF_UP (padrão jurídico-contábil RFB/BCB), 2 casas decimais nos valores monetários.")
    L.append("")
    L.append("---")
    L.append("")
    L.append("## 2. Universo de Faturas")
    L.append("")
    L.append(f"Total de **{totais['n_total']} faturas** com cadeia documental no SPCF da PRODAM, divididas em:")
    L.append("")
    L.append(f"- **Tier 1 ({totais['n_tier1']} faturas)** — vencimento ainda dentro do prazo prescricional quinquenal (Art. 206 §5º I CC).")
    L.append(f"- **Tier 2 ({totais['n_tier2']} faturas)** — ressuscitadas por Nota de Empenho de 2026 nos termos do **Art. 202 VI CC** (reconhecimento tácito interrompe prescrição, reiniciando-a pela metade na forma do **Decreto 20.910/1932** = 2,5 anos a partir da nova NE).")
    L.append("")
    L.append("Excluída a fatura **id 122703** (R$ 29.001,42) por prescrição consumada (D-71 em relação à data-base).")
    L.append("")
    L.append("---")
    L.append("")
    L.append("## 3. Memorial Fatura-a-Fatura")
    L.append("")
    L.append("| # | ID | NF | Contrato | Comp. | Venc. | Tier | Valor Bruto | Meses | Fator SELIC | Valor Atualizado |")
    L.append("|---|----|----|----------|-------|-------|------|-------------:|------:|------------:|------------------:|")
    for i, r in enumerate(sorted(resultados, key=lambda x: (x["tier"], x["contrato"], x["vencimento"])), 1):
        L.append(
            f"| {i} | {r['id']} | {r['nf']} | {r['contrato']} | "
            f"{r['competencia'].strftime('%m/%Y')} | {r['vencimento'].strftime('%d/%m/%Y')} | "
            f"{r['tier']} | R$ {fmt_brl(r['valor_bruto'])} | {r['meses_entre']} | "
            f"{r['fator_selic']:.6f} | R$ {fmt_brl(r['valor_atualizado'])} |"
        )
    L.append("")
    L.append("---")
    L.append("")
    L.append("## 4. Totais Consolidados")
    L.append("")
    L.append("| Métrica | Valor |")
    L.append("|---------|------:|")
    L.append(f"| Faturas Tier 1 (prazo natural 5a) | {totais['n_tier1']} |")
    L.append(f"| Faturas Tier 2 (ressuscitadas NE 2026) | {totais['n_tier2']} |")
    L.append(f"| **Principal bruto Tier 1** | R$ {fmt_brl(totais['tier1_principal_bruto'])} |")
    L.append(f"| **Principal bruto Tier 2** | R$ {fmt_brl(totais['tier2_principal_bruto'])} |")
    L.append(f"| **Principal bruto TOTAL** | **R$ {fmt_brl(totais['principal_bruto'])}** |")
    L.append(f"| Correção + juros SELIC | R$ {fmt_brl(totais['correcao_juros_selic'])} |")
    L.append(f"| Fator SELIC médio do conjunto | {totais['fator_medio']:.6f}x |")
    L.append(f"| **VALOR ATUALIZADO TOTAL** | **R$ {fmt_brl(totais['total_atualizado'])}** |")
    L.append(f"| Honorários PRODAM (20%) | R$ {fmt_brl(totais['honorarios_20pct'])} |")
    L.append("")
    L.append("---")
    L.append("")
    L.append("## 5. Caráter Preliminar do Memorial")
    L.append("")
    L.append("Este memorial é **preliminar** para os fins exclusivos da **Termo de Reconhecimento de Dívida (TRD)**, em conformidade com o regime SELIC pura aplicável à administração direta estadual por norma legal cogente (EC 113/2021).")
    L.append("")
    L.append("**Não substitui memorial pericial detalhado** que será produzido, se necessário, em fase de **execução judicial** ou **ação monitória**, oportunidade em que será realizada extração formal de cláusulas contratuais (regime, índice, juros, multa) de cada um dos 13 contratos identificados no SPCF, via skill `extracao-clausulas-contratuais` do escritório.")
    L.append("")
    L.append("**Reserva**: caso a SES/SUSAM, em resposta à TRD, alegue regime contratual específico (ex.: cláusula que vincule a contrato com índice diverso), o memorial será refeito por contrato com a metodologia adequada e os valores recalculados antes de qualquer ajuizamento.")
    L.append("")
    L.append("---")
    L.append("")
    L.append("Manaus/AM, " + DATA_BASE.strftime("%d/%m/%Y") + ".")
    L.append("")
    L.append("**Gabriel Mar** — OAB/AM 15.697  ")
    L.append("Gabriel Mar Sociedade Individual de Advocacia")
    L.append("")
    return "\n".join(L)

# ─── Render XLSX ───────────────────────────────────────────────────────────
def render_xlsx(resultados: list[dict], totais: dict, out_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Faturas"
    headers = ["#", "ID", "NF", "Contrato", "Competência", "Vencimento",
               "Tier", "Valor Bruto", "Meses", "Fator SELIC",
               "Correção+Juros SELIC", "Valor Atualizado", "Última NE Pós-Venc"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="2D2D2D")
        cell.font = Font(bold=True, color="FFFFFF")

    for i, r in enumerate(sorted(resultados, key=lambda x: (x["tier"], x["contrato"], x["vencimento"])), 1):
        ws.append([
            i, r["id"], r["nf"], r["contrato"],
            r["competencia"].strftime("%m/%Y"),
            r["vencimento"].isoformat(),
            r["tier"],
            float(r["valor_bruto"]), r["meses_entre"], float(r["fator_selic"]),
            float(r["correcao_juros_selic"]), float(r["valor_atualizado"]),
            r["ultima_ne_pos_venc"] or "",
        ])

    ws2 = wb.create_sheet("Por_Contrato")
    ws2.append(["Contrato", "Tier dominante", "Faturas", "Principal Bruto",
                "Correção SELIC", "Total Atualizado"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    contratos = sorted({r["contrato"] for r in resultados})
    for c in contratos:
        sub = [r for r in resultados if r["contrato"] == c]
        tier_dom = max(set(r["tier"] for r in sub), key=lambda t: sum(1 for r in sub if r["tier"] == t))
        principal = sum((r["valor_bruto"] for r in sub), ZERO)
        correcao = sum((r["correcao_juros_selic"] for r in sub), ZERO)
        total = sum((r["valor_atualizado"] for r in sub), ZERO)
        ws2.append([c, f"Tier {tier_dom}", len(sub),
                    float(principal), float(correcao), float(total)])

    ws3 = wb.create_sheet("Metadados")
    ws3.append(["Chave", "Valor"])
    for cell in ws3[1]:
        cell.font = Font(bold=True)
    meta = [
        ("Devedor", "SES/SUSAM"),
        ("CNPJ", "00.697.295/0001-05"),
        ("Data-base", DATA_BASE.isoformat()),
        ("Faturas candidatas TRD", totais["n_total"]),
        ("Tier 1 (prazo natural 5a)", totais["n_tier1"]),
        ("Tier 2 (ressuscitadas NE 2026)", totais["n_tier2"]),
        ("Principal bruto total", float(totais["principal_bruto"])),
        ("Correção+juros SELIC", float(totais["correcao_juros_selic"])),
        ("Total atualizado", float(totais["total_atualizado"])),
        ("Honorários PRODAM (20%)", float(totais["honorarios_20pct"])),
        ("Fator SELIC médio", float(totais["fator_medio"])),
        ("Série BCB", 4390),
        ("Regime aplicado", "SELIC pura (EC 113/2021 Art. 3º)"),
        ("Caráter", "PRELIMINAR — TRD, não pericial"),
        ("Excluída prescrita", "id 122703 (D-71)"),
        ("Gerado em", dt.datetime.now().isoformat(timespec="seconds")),
        ("Script", "scripts/ad_hoc/gerar_memorial_preliminar_ses.py"),
    ]
    for k, v in meta:
        ws3.append([k, v])
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

# ─── Main ──────────────────────────────────────────────────────────────────
def main() -> int:
    print("=== Memorial Preliminar SES/SUSAM ===")
    print(f"DB: {DB_PATH}")
    print(f"Output: {OUT_DIR}")
    print(f"Data-base: {DATA_BASE.isoformat()}")
    print()
    candidatas = selecionar_candidatas()
    print(f"Candidatas a TRD selecionadas: {len(candidatas)}")
    tier1 = [c for c in candidatas if c["tier"] == "1"]
    tier2 = [c for c in candidatas if c["tier"] == "2"]
    print(f"  Tier 1: {len(tier1)} ({sum((c['valor_bruto'] for c in tier1), ZERO)})")
    print(f"  Tier 2: {len(tier2)} ({sum((c['valor_bruto'] for c in tier2), ZERO)})")
    print()
    resultados, totais = calcular_memorial(candidatas)
    print(f"Total bruto: R$ {fmt_brl(totais['principal_bruto'])}")
    print(f"Total atualizado: R$ {fmt_brl(totais['total_atualizado'])}")
    print(f"Fator SELIC médio: {totais['fator_medio']:.6f}x")
    print()
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    md_path = OUT_DIR / f"MEMORIAL_PRELIMINAR_SES_SUSAM_{DATA_BASE.isoformat()}.md"
    xlsx_path = OUT_DIR / f"MEMORIAL_PRELIMINAR_SES_SUSAM_{DATA_BASE.isoformat()}.xlsx"
    md_path.write_text("﻿" + render_md(resultados, totais), encoding="utf-8")
    render_xlsx(resultados, totais, xlsx_path)
    print(f"Memorial MD:   {md_path}")
    print(f"Memorial XLSX: {xlsx_path}")
    return 0

if __name__ == "__main__":
    sys.exit(main())
