"""
Comparativo Tema 810 (STF) vs Lei 14.905/24 — 26 faturas DETRAN silentes
=========================================================================

Cenário A — STF Tema 810:
  - Vencimento -> 29/08/2024: IPCA-E (16121) + Poupança PF (25)
  - 30/08/2024 -> 02/05/2026: SELIC pura (4390)

Cenário B — INPC + SELIC (Lei 14.905/24 desde o vencimento):
  - Vencimento -> 02/05/2026: INPC (188) + SELIC (4390)

Critério metodológico (Manual CJF):
  - Índices/taxas mensais aplicados de forma cumulativa: produto de (1 + var/100)
  - Janela: do mês do vencimento (inclusive) até o mês da data-base (inclusive)
  - Para o corte 30/08/2024, agosto/2024 entra no Cenário A regime IPCA+Poupança
    (último mês do regime antigo) e setembro/2024 inicia regime SELIC pura.
  - Decimal em todos os cálculos monetários.
"""

from __future__ import annotations

import csv
import json
import time
import urllib.request
from collections import OrderedDict
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_EVEN, getcontext
from pathlib import Path

getcontext().prec = 28

BASE = Path(r"C:\Users\gabri\Desktop\PROJETO_PRODAM\DETRAN_AUDITORIA_COMPLETA")
CSV_PATH = BASE / "88_FATURAS_RECONCILIAR.csv"
OUT_DIR = BASE / "COMPARATIVO_TEMA810_vs_LEI14905"
OUT_DIR.mkdir(parents=True, exist_ok=True)

DATA_BASE = date(2026, 5, 2)
CORTE_TEMA810 = date(2024, 8, 29)  # último dia do regime IPCA+Poupança
REGIME_LABEL = "INPC + SELIC (Lei 14.905/24)"

SERIES = OrderedDict(
    [
        ("IPCA_E", 16121),  # IPCA-15 (variação % mensal)
        ("INPC", 188),  # INPC (variação % mensal)
        ("SELIC", 4390),  # SELIC acumulada no mês (% a.m.)
        (
            "POUPANCA",
            25,
        ),  # Rendimento poupança PF (% a.m. — já consolida regra Lei 12.703/2012)
    ]
)

DATA_INICIO_BCB = "01/01/2019"
DATA_FIM_BCB = "02/05/2026"


def fetch_bcb_serie(codigo: int, max_retries: int = 4) -> list[dict]:
    """Baixa série temporal do BCB no formato JSON, com retry."""
    url = (
        f"https://api.bcb.gov.br/dados/serie/bcdata.sgs.{codigo}/dados"
        f"?formato=json&dataInicial={DATA_INICIO_BCB}&dataFinal={DATA_FIM_BCB}"
    )
    ultimo_erro = None
    for tentativa in range(1, max_retries + 1):
        try:
            req = urllib.request.Request(
                url, headers={"User-Agent": "Mozilla/5.0 (PRODAM-calculo-juridico)"}
            )
            with urllib.request.urlopen(req, timeout=60) as resp:
                body = resp.read().decode("utf-8")
                if not body.strip().startswith("["):
                    raise ValueError(
                        f"Resposta nao-JSON (primeiros 200 chars): {body[:200]!r}"
                    )
                return json.loads(body)
        except Exception as e:
            ultimo_erro = e
            espera = 2 * tentativa
            print(
                f"    [retry {tentativa}/{max_retries}] {type(e).__name__}: {e} -- aguardando {espera}s"
            )
            time.sleep(espera)
    raise RuntimeError(
        f"Falha ao baixar serie {codigo} apos {max_retries} tentativas: {ultimo_erro}"
    )


def baixar_todas_series() -> dict[str, dict[tuple[int, int], Decimal]]:
    """Retorna dict {nome_serie: {(ano, mes): valor_decimal}}."""
    resultado: dict[str, dict[tuple[int, int], Decimal]] = {}
    for nome, codigo in SERIES.items():
        print(f"  Baixando série {nome} (BCB {codigo})...", flush=True)
        dados = fetch_bcb_serie(codigo)
        if not dados:
            raise RuntimeError(f"Série {nome} ({codigo}) retornou vazio. PARANDO.")
        mensal: dict[tuple[int, int], Decimal] = {}
        for item in dados:
            d = datetime.strptime(item["data"], "%d/%m/%Y").date()
            valor = Decimal(item["valor"])
            mensal[(d.year, d.month)] = valor
        resultado[nome] = mensal
        time.sleep(1.5)  # rate limit
    return resultado


def meses_no_intervalo(inicio: date, fim: date) -> list[tuple[int, int]]:
    """Lista de (ano, mês) entre inicio e fim, inclusive."""
    if fim < inicio:
        return []
    meses = []
    y, m = inicio.year, inicio.month
    while (y, m) <= (fim.year, fim.month):
        meses.append((y, m))
        m += 1
        if m > 12:
            m = 1
            y += 1
    return meses


def fator_acumulado(
    serie: dict[tuple[int, int], Decimal], meses: list[tuple[int, int]], nome_serie: str
) -> tuple[Decimal, list[dict]]:
    """Multiplica (1 + var/100) ao longo dos meses. Retorna (fator, trilha_auditoria)."""
    fator = Decimal("1")
    trilha = []
    for ym in meses:
        if ym not in serie:
            raise RuntimeError(
                f"Série {nome_serie} sem dado para {ym[1]:02d}/{ym[0]}. PARANDO."
            )
        var_pct = serie[ym]
        f_mes = Decimal("1") + var_pct / Decimal("100")
        fator *= f_mes
        trilha.append(
            {
                "mes": f"{ym[1]:02d}/{ym[0]}",
                "var_pct": str(var_pct),
                "fator_mes": str(f_mes),
            }
        )
    return fator, trilha


def calcular_cenario_A(
    valor_orig: Decimal, vencimento: date, series: dict
) -> tuple[Decimal, dict]:
    """Cenário A — Tema 810."""
    detalhe = {}

    # Sub-período 1: vencimento -> agosto/2024 (inclusive)
    if vencimento <= CORTE_TEMA810:
        meses_p1 = meses_no_intervalo(vencimento, CORTE_TEMA810)
        f_ipca, t_ipca = fator_acumulado(series["IPCA_E"], meses_p1, "IPCA_E")
        f_poup, t_poup = fator_acumulado(series["POUPANCA"], meses_p1, "POUPANCA")
        # IPCA-E é correção, Poupança é juros — ambos aplicados sobre o valor original
        # Critério Manual CJF: VC = V0 * fator_correcao; Juros = V0 * (fator_juros - 1)
        # Total ao final do P1 = V0 * fator_correcao + V0 * (fator_juros - 1)
        # Simplificação consagrada: Total_P1 = V0 * (f_correcao + f_juros - 1)
        # (juros simples sobre o valor original, conforme jurisprudência STF Tema 810)
        valor_pos_p1 = valor_orig * (f_ipca + f_poup - Decimal("1"))
        detalhe["periodo_1"] = {
            "regime": "IPCA-E (correção) + Poupança PF (juros) — Manual CJF / STF Tema 810",
            "data_inicio": vencimento.isoformat(),
            "data_fim": CORTE_TEMA810.isoformat(),
            "fator_ipca_e": str(f_ipca),
            "fator_poupanca": str(f_poup),
            "valor_apos_p1": str(valor_pos_p1),
            "trilha_ipca": t_ipca,
            "trilha_poupanca": t_poup,
        }
    else:
        # Vencimento posterior ao corte — pula direto para SELIC
        valor_pos_p1 = valor_orig
        detalhe["periodo_1"] = {
            "regime": "N/A — vencimento posterior ao corte de 29/08/2024",
            "valor_apos_p1": str(valor_orig),
        }

    # Sub-período 2: 01/09/2024 (ou vencimento se posterior) -> data-base
    inicio_p2 = max(date(2024, 9, 1), date(vencimento.year, vencimento.month, 1))
    if vencimento > CORTE_TEMA810:
        inicio_p2 = vencimento
    meses_p2 = meses_no_intervalo(inicio_p2, DATA_BASE)
    if meses_p2:
        f_selic, t_selic = fator_acumulado(series["SELIC"], meses_p2, "SELIC")
        # SELIC pura SUBSTITUI correção+juros. Aplica sobre o valor já corrigido em P1.
        # Critério: SELIC é taxa composta no Manual CJF (art. 406 CC)
        valor_final = valor_pos_p1 * f_selic
        detalhe["periodo_2"] = {
            "regime": "SELIC pura (substitui correção+juros)",
            "data_inicio": inicio_p2.isoformat(),
            "data_fim": DATA_BASE.isoformat(),
            "fator_selic": str(f_selic),
            "valor_final": str(valor_final),
            "trilha_selic": t_selic,
        }
    else:
        valor_final = valor_pos_p1
        detalhe["periodo_2"] = {"regime": "N/A", "valor_final": str(valor_pos_p1)}

    return valor_final, detalhe


def calcular_cenario_B(
    valor_orig: Decimal, vencimento: date, series: dict
) -> tuple[Decimal, dict]:
    """Cenário B — INPC + SELIC desde o vencimento."""
    meses = meses_no_intervalo(vencimento, DATA_BASE)
    f_inpc, t_inpc = fator_acumulado(series["INPC"], meses, "INPC")
    f_selic, t_selic = fator_acumulado(series["SELIC"], meses, "SELIC")
    # Cenário B: INPC corrige, SELIC é juros. Ambos sobre o valor original
    # (Manual CJF para juros simples sobre principal corrigido seria diferente,
    # mas aqui aplicamos a regra usual da Lei 14.905/24: ambos somados como
    # incremento sobre V0).
    # Total = V0 * (f_inpc + f_selic - 1)
    valor_final = valor_orig * (f_inpc + f_selic - Decimal("1"))
    detalhe = {
        "regime": "INPC (correção) + SELIC (juros) — Lei 14.905/24",
        "data_inicio": vencimento.isoformat(),
        "data_fim": DATA_BASE.isoformat(),
        "fator_inpc": str(f_inpc),
        "fator_selic": str(f_selic),
        "valor_final": str(valor_final),
        "trilha_inpc": t_inpc,
        "trilha_selic": t_selic,
    }
    return valor_final, detalhe


def quant_brl(d: Decimal) -> Decimal:
    return d.quantize(Decimal("0.01"), rounding=ROUND_HALF_EVEN)


def fmt_brl(d: Decimal) -> str:
    s = f"{quant_brl(d):,.2f}"
    return "R$ " + s.replace(",", "X").replace(".", ",").replace("X", ".")


def main():
    print("=" * 70)
    print("COMPARATIVO TEMA 810 (STF) vs LEI 14.905/24 — 26 faturas DETRAN")
    print("=" * 70)

    # 1) Filtrar CSV
    print("\n[1/4] Lendo CSV e filtrando faturas...")
    faturas = []
    with open(CSV_PATH, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f, delimiter=";")
        for row in reader:
            if row["regime_label"].strip() == REGIME_LABEL:
                faturas.append(row)
    print(f"  Filtradas: {len(faturas)} faturas")

    # 2) Baixar séries BCB
    print("\n[2/4] Baixando séries do BCB...")
    series = baixar_todas_series()
    for nome, dados in series.items():
        amostras = sorted(dados.keys())
        print(f"  {nome}: {len(dados)} pontos | {amostras[0]} -> {amostras[-1]}")

    # 3) Calcular cenários
    print("\n[3/4] Calculando cenários A e B para cada fatura...")
    resultados = []
    memoria = {
        "metadata": {
            "data_base": DATA_BASE.isoformat(),
            "corte_tema_810": CORTE_TEMA810.isoformat(),
            "fonte_series": {nome: f"BCB SGS {cod}" for nome, cod in SERIES.items()},
            "criterio_metodologico": (
                "Manual CJF: fator acumulado = produto de (1 + var_pct/100) "
                "sobre meses no intervalo [mes_vencimento, mes_data_base]. "
                "Cenário A: V0 * (f_ipca + f_poup - 1) até 29/08/2024, "
                "depois SELIC composta sobre o resultado. "
                "Cenário B: V0 * (f_inpc + f_selic - 1) desde o vencimento."
            ),
            "decimal_precision": 28,
            "data_calculo": datetime.now().isoformat(),
        },
        "series_bcb": {},
        "faturas": [],
    }

    # Memorial das séries (compactado: só pontos no range das faturas)
    todos_meses_usados = set()
    for f in faturas:
        venc = datetime.strptime(f["data_vencimento"], "%d/%m/%Y").date()
        for ym in meses_no_intervalo(venc, DATA_BASE):
            todos_meses_usados.add(ym)

    for nome, dados in series.items():
        memoria["series_bcb"][nome] = {
            "codigo_bcb": SERIES[nome],
            "qtd_pontos_total": len(dados),
            "valores_por_mes": {
                f"{ym[0]}-{ym[1]:02d}": str(dados[ym])
                for ym in sorted(todos_meses_usados)
                if ym in dados
            },
        }

    for f in faturas:
        venc = datetime.strptime(f["data_vencimento"], "%d/%m/%Y").date()
        v0 = Decimal(f["valor_original"])

        v_a, det_a = calcular_cenario_A(v0, venc, series)
        v_b, det_b = calcular_cenario_B(v0, venc, series)
        diff = v_b - v_a
        diff_pct = (diff / v_a * Decimal("100")) if v_a != 0 else Decimal("0")

        resultados.append(
            {
                "contrato": f["contrato"],
                "fatura": f["numero_fatura"],
                "vencimento": venc,
                "valor_original": quant_brl(v0),
                "valor_cenario_A": quant_brl(v_a),
                "valor_cenario_B": quant_brl(v_b),
                "diferenca_BRL": quant_brl(diff),
                "diferenca_pct": diff_pct.quantize(
                    Decimal("0.01"), rounding=ROUND_HALF_EVEN
                ),
            }
        )

        memoria["faturas"].append(
            {
                "contrato": f["contrato"],
                "fatura": f["numero_fatura"],
                "vencimento": venc.isoformat(),
                "valor_original": str(v0),
                "cenario_A": {
                    "valor_final": str(quant_brl(v_a)),
                    "detalhe": det_a,
                },
                "cenario_B": {
                    "valor_final": str(quant_brl(v_b)),
                    "detalhe": det_b,
                },
            }
        )

    # 4) Gerar entregáveis
    print("\n[4/4] Gerando entregáveis...")

    # 4a) XLSX
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  Instalando openpyxl...")
        import subprocess

        subprocess.check_call(["py", "-3.12", "-m", "pip", "install", "openpyxl", "-q"])
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Comparativo"

    headers = [
        "Contrato",
        "Fatura",
        "Vencimento",
        "Valor Original (R$)",
        "Cenário A — Tema 810 (R$)",
        "Cenário B — Lei 14.905/24 (R$)",
        "Diferença B-A (R$)",
        "Diferença B-A (%)",
    ]
    ws.append(headers)
    for cell in ws[1]:/
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(
            start_color="1F4E78", end_color="1F4E78", fill_type="solid"
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in resultados:
        ws.append(
            [
                r["contrato"],
                r["fatura"],
                r["vencimento"].strftime("%d/%m/%Y"),
                float(r["valor_original"]),
                float(r["valor_cenario_A"]),
                float(r["valor_cenario_B"]),
                float(r["diferenca_BRL"]),
                float(r["diferenca_pct"]),
            ]
        )

    # Linha de totais
    total_v0 = sum(r["valor_original"] for r in resultados)
    total_a = sum(r["valor_cenario_A"] for r in resultados)
    total_b = sum(r["valor_cenario_B"] for r in resultados)
    total_diff = total_b - total_a
    total_diff_pct = (
        (total_diff / total_a * Decimal("100")) if total_a != 0 else Decimal("0")
    )

    ws.append([])
    ws.append(
        [
            "TOTAL",
            "",
            "",
            float(total_v0),
            float(total_a),
            float(total_b),
            float(total_diff),
            float(total_diff_pct.quantize(Decimal("0.01"))),
        ]
    )
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
        )

    # Formatação de colunas monetárias
    for col_idx in [4, 5, 6, 7]:
        col_letter = get_column_letter(col_idx)
        for row in range(2, ws.max_row + 1):
            ws[f"{col_letter}{row}"].number_format = "#,##0.00"
    for row in range(2, ws.max_row + 1):
        ws[f"H{row}"].number_format = '0.00"%"'

    # Auto-width
    for col in ws.columns:
        max_len = max(
            (len(str(c.value)) for c in col if c.value is not None), default=10
        )
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 32)

    xlsx_path = OUT_DIR / "comparativo.xlsx"
    wb.save(xlsx_path)
    print(f"  [OK] {xlsx_path}")

    # 4b) JSON memoria
    json_path = OUT_DIR / "memoria_calculo.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(memoria, f, ensure_ascii=False, indent=2)
    print(f"  [OK] {json_path}")

    # 4c) Resumo MD
    md_path = OUT_DIR / "resumo.md"
    diferenca_label = "FAVORÁVEL CENÁRIO B" if total_diff > 0 else "FAVORÁVEL CENÁRIO A"
    cenario_maior = "B" if total_b > total_a else "A"
    cenario_menor = "A" if total_b > total_a else "B"
    valor_maior = total_b if total_b > total_a else total_a
    valor_menor = total_a if total_b > total_a else total_b

    venc_min = min(r["vencimento"] for r in resultados)
    venc_max = max(r["vencimento"] for r in resultados)

    with open(md_path, "w", encoding="utf-8") as f:
        f.write(f"""# Comparativo: STF Tema 810 vs Lei 14.905/24

**Data-base do cálculo**: {DATA_BASE.strftime('%d/%m/%Y')}
**Faturas analisadas**: {len(resultados)} (regime "{REGIME_LABEL}")
**Vencimentos**: {venc_min.strftime('%d/%m/%Y')} a {venc_max.strftime('%d/%m/%Y')}

---

## Totais

| Métrica | Valor |
|---|---|
| Valor original (somatório) | **{fmt_brl(total_v0)}** |
| Cenário A (Tema 810: IPCA-E + Poupança até 29/08/2024, SELIC depois) | **{fmt_brl(total_a)}** |
| Cenário B (Lei 14.905/24: INPC + SELIC desde o vencimento) | **{fmt_brl(total_b)}** |
| Diferença (B − A) | **{fmt_brl(total_diff)}** |
| Diferença % (sobre A) | **{total_diff_pct.quantize(Decimal('0.01'))}%** |

---

## Recomendação numérica

O Cenário **{cenario_maior}** produz o **maior valor atualizado**: {fmt_brl(valor_maior)}.
O Cenário **{cenario_menor}** produz {fmt_brl(valor_menor)}.

A diferença absoluta é de **{fmt_brl(abs(total_diff))}** ({abs(total_diff_pct).quantize(Decimal('0.01'))}% sobre o cenário menor).

**{diferenca_label}** — do ponto de vista exclusivamente numérico, defender o Cenário **{cenario_maior}** maximiza o valor a recuperar para essas 26 faturas.

> Esta recomendação é **estritamente quantitativa**. A escolha do regime aplicável depende de análise jurídica adicional (qualificação do devedor como Fazenda Pública, data dos fatos, posição do STJ/STF sobre a aplicabilidade temporal da Lei 14.905/24 a obrigações pré-existentes).

---

## Séries do BCB utilizadas

| Série | Código BCB | Uso | Pontos baixados |
|---|---|---|---|
| IPCA-15 (proxy IPCA-E) | {SERIES['IPCA_E']} | Cenário A — correção até 29/08/2024 | {len(series['IPCA_E'])} |
| INPC | {SERIES['INPC']} | Cenário B — correção desde vencimento | {len(series['INPC'])} |
| SELIC acumulada no mês | {SERIES['SELIC']} | A: substitui após 29/08/2024 / B: juros | {len(series['SELIC'])} |
| Rendimento poupança PF | {SERIES['POUPANCA']} | Cenário A — juros até 29/08/2024 (Lei 12.703/2012) | {len(series['POUPANCA'])} |

---

## Critério metodológico

- Decimal (`decimal.Decimal`, precisão 28) em todos os cálculos monetários
- Fator acumulado = ∏ (1 + var_pct/100) sobre os meses no intervalo [mês_vencimento, mês_data_base]
- Cenário A — período 1 (até 29/08/2024): V₀ × (f_IPCA + f_Poupança − 1)
- Cenário A — período 2 (após 30/08/2024): valor_p1 × f_SELIC (SELIC composta substituindo correção+juros)
- Cenário B: V₀ × (f_INPC + f_SELIC − 1) desde o vencimento até 02/05/2026
- Critério juros do Cenário A consistente com Manual CJF (Justiça Federal): juros simples sobre o valor original na fase pré-Tema 810
- Não aplicada multa (2%) nem juros de mora (1% a.m.) — esses são exclusivos do CT 22/2014, fora do escopo das 26 faturas em regime Lei 14.905/24

---

## Auditoria

Detalhamento mês-a-mês (todos os fatores de cada série usados em cada fatura) está em `memoria_calculo.json`.
""")
    print(f"  [OK] {md_path}")

    # Sumário no terminal
    print("\n" + "=" * 70)
    print("RESUMO FINAL")
    print("=" * 70)
    print(f"  Faturas:                {len(resultados)}")
    print(f"  Valor original total:   {fmt_brl(total_v0)}")
    print(f"  Cenário A (Tema 810):   {fmt_brl(total_a)}")
    print(f"  Cenário B (Lei 14.905): {fmt_brl(total_b)}")
    print(
        f"  Diferença (B - A):      {fmt_brl(total_diff)}  ({total_diff_pct.quantize(Decimal('0.01'))}%)"
    )
    print(f"  Cenário mais favorável: {cenario_maior}")
    print("=" * 70)


if __name__ == "__main__":
    main()
